"""
Ombor — qoldiqlar, eksport/import, ombordan omborga o'tkazish.
"""
import io
import traceback
from datetime import datetime
from urllib.parse import quote, unquote

import openpyxl
from fastapi import APIRouter, Request, Depends, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, and_, func, text

from app.core import templates
from app.models.database import (
    get_db,
    User,
    Warehouse,
    Stock,
    Product,
    Purchase,
    PurchaseItem,
    Production,
    Recipe,
    StockAdjustmentDoc,
    StockAdjustmentDocItem,
    WarehouseTransfer,
    WarehouseTransferItem,
)
from app.deps import require_auth, require_admin
from app.utils.user_scope import get_warehouses_for_user

router = APIRouter(prefix="/warehouse", tags=["warehouse"])


def _warehouses_for_user(db: Session, user: User):
    """Foydalanuvchi uchun ko'rinadigan omborlar: sozlamada belgilangan yoki admin/raxbar uchun barcha."""
    return get_warehouses_for_user(db, user)


@router.get("", response_class=HTMLResponse)
async def warehouse_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    warehouses = _warehouses_for_user(db, current_user)
    wh_ids = [w.id for w in warehouses]
    stocks_q = db.query(Stock).join(Product).join(Warehouse).filter(Stock.quantity > 0)
    if wh_ids:
        stocks_q = stocks_q.filter(Stock.warehouse_id.in_(wh_ids))
    stocks = stocks_q.all()
    stock_sources = {}
    for s in stocks:
        items = []
        purchases = (
            db.query(Purchase)
            .join(PurchaseItem, Purchase.id == PurchaseItem.purchase_id)
            .filter(
                Purchase.warehouse_id == s.warehouse_id,
                PurchaseItem.product_id == s.product_id,
                Purchase.status == "confirmed",
            )
            .order_by(Purchase.date.desc())
            .limit(3)
            .all()
        )
        for p in purchases:
            items.append((p.number, f"/purchases/edit/{p.id}", p.date.strftime("%d.%m.%Y") if p.date else ""))
        out_wh_id = s.warehouse_id
        # Production query - max_stage muammosini oldini olish
        # Bazada max_stage ustuni bo'lmasligi mumkin, shuning uchun faqat mavjud ustunlarni ishlatamiz
        try:
            # Faqat kerakli ustunlarni tanlab, max_stage ni o'z ichiga olmaydigan so'rov
            # SQLAlchemy modelda max_stage bor, lekin bazada yo'q bo'lishi mumkin
            result = db.execute(
                text("""
                    SELECT p.id, p.number, p.date
                    FROM productions p
                    INNER JOIN recipes r ON p.recipe_id = r.id
                    WHERE p.status = 'completed'
                      AND r.product_id = :product_id
                      AND (
                          p.output_warehouse_id = :warehouse_id
                          OR (p.output_warehouse_id IS NULL AND p.warehouse_id = :warehouse_id)
                      )
                    ORDER BY p.date DESC
                    LIMIT 3
                """),
                {"product_id": s.product_id, "warehouse_id": out_wh_id}
            )
            productions = result.fetchall()
            for pr in productions:
                pr_date = pr.date.strftime("%d.%m.%Y") if pr.date else ""
                items.append((pr.number, "/production/orders", pr_date))
        except Exception as prod_error:
            # Database da max_stage yoki boshqa ustunlar yo'q bo'lishi mumkin - e'tiborsiz qoldiramiz
            print(f"Production query error (warehouse {s.warehouse_id}, product {s.product_id}): {prod_error}")
            import traceback
            traceback.print_exc()
        adj_docs = (
            db.query(StockAdjustmentDoc)
            .join(StockAdjustmentDocItem, StockAdjustmentDoc.id == StockAdjustmentDocItem.doc_id)
            .filter(
                StockAdjustmentDoc.status == "confirmed",
                StockAdjustmentDocItem.warehouse_id == s.warehouse_id,
                StockAdjustmentDocItem.product_id == s.product_id,
            )
            .order_by(StockAdjustmentDoc.date.desc())
            .limit(3)
            .distinct()
            .all()
        )
        for doc in adj_docs:
            items.append((doc.number, f"/qoldiqlar/tovar/hujjat/{doc.id}", doc.date.strftime("%d.%m.%Y") if doc.date else ""))
        items.sort(key=lambda x: x[2] or "", reverse=True)
        stock_sources[s.id] = items[:8]
    # Hujjatlar ro'yxati — mahsulotlar hujjat ichida ko'riladi
    qoldiq_docs = (
        db.query(StockAdjustmentDoc)
        .order_by(StockAdjustmentDoc.date.desc(), StockAdjustmentDoc.id.desc())
        .limit(100)
        .all()
    )
    purchase_docs = (
        db.query(Purchase)
        .filter(Purchase.status.in_(["confirmed", "draft"]))
        .order_by(Purchase.date.desc())
        .limit(80)
        .all()
    )
    return templates.TemplateResponse("warehouse/list.html", {
        "request": request,
        "warehouses": warehouses,
        "stocks": stocks,
        "stock_sources": stock_sources,
        "qoldiq_docs": qoldiq_docs,
        "purchase_docs": purchase_docs,
        "current_user": current_user,
        "page_title": "Ombor qoldiqlari",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


@router.post("/stock/{stock_id}/zero")
async def warehouse_stock_zero(
    stock_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    stock = db.query(Stock).filter(Stock.id == stock_id).first()
    if not stock:
        raise HTTPException(status_code=404, detail="Qoldiq topilmadi")
    stock.quantity = 0
    db.commit()
    return RedirectResponse(url="/warehouse", status_code=303)


@router.get("/export")
async def warehouse_export(db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    stocks = db.query(Stock).join(Product).join(Warehouse).filter(Stock.quantity > 0).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qoldiqlar"
    ws.append(["Ombor nomi", "Ombor kodi", "Mahsulot kodi", "Mahsulot nomi", "Qoldiq", "Tannarx (so'm)", "Summa (so'm)"])
    for s in stocks:
        pr, wh = s.product, s.warehouse
        tannarx = (pr.purchase_price or 0) if pr else 0
        summa = s.quantity * tannarx
        ws.append([
            wh.name if wh else "",
            wh.code if wh else "",
            pr.code if pr else "",
            pr.name if pr else "",
            s.quantity,
            tannarx,
            summa,
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ombor_qoldiqlari.xlsx"},
    )


@router.get("/template")
async def warehouse_template(db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Andoza"
    ws.append(["Ombor nomi (yoki kodi)", "Mahsulot nomi (yoki kodi)", "Qoldiq", "Tannarx (so'm)", "Sotuv narxi (so'm)"])
    ws.append(["Xom ashyo ombori", "Yong'oq", 30, "", ""])
    ws.append(["Xom ashyo ombori", "Bodom", 100, "", ""])
    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=qoldiqlar_andoza.xlsx"},
    )


@router.post("/import")
async def warehouse_import(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    form = await request.form()
    file = form.get("file") or form.get("excel_file")
    if not file or not getattr(file, "filename", None):
        return RedirectResponse(url="/warehouse?error=import&detail=" + quote("Excel fayl tanlang"), status_code=303)
    try:
        contents = await file.read()
        if not contents:
            return RedirectResponse(url="/warehouse?error=import&detail=" + quote("Fayl bo'sh"), status_code=303)
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=False, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        updated, skipped = 0, 0
        missing_products, missing_warehouses = [], []
        for row in rows:
            if not row or (row[0] is None and row[1] is None):
                continue
            wh_key = str(row[0] or "").strip() if len(row) > 0 else ""
            raw_prod = row[1] if len(row) > 1 else None
            prod_key = str(int(float(raw_prod))) if raw_prod is not None and isinstance(raw_prod, (int, float)) and float(raw_prod) == int(float(raw_prod)) else str(raw_prod or "").strip()
            try:
                qty = float(row[2]) if len(row) > 2 and row[2] is not None else 0
            except (TypeError, ValueError):
                qty = 0
            tannarx = None
            sotuv_narxi = None
            if len(row) > 3 and row[3] is not None and row[3] != "":
                try:
                    tannarx = float(row[3])
                except (TypeError, ValueError):
                    pass
            if len(row) > 4 and row[4] is not None and row[4] != "":
                try:
                    sotuv_narxi = float(row[4])
                except (TypeError, ValueError):
                    pass
            if not wh_key or not prod_key:
                skipped += 1
                continue
            warehouse = db.query(Warehouse).filter(
                (func.lower(Warehouse.name) == wh_key.lower()) | (Warehouse.code == wh_key)
            ).first()
            # Kod/shtrixkod va nom — katta-kichik harf farqisiz
            product = db.query(Product).filter(
                or_(
                    and_(Product.code.isnot(None), Product.code != "", func.lower(Product.code) == prod_key.lower()),
                    and_(Product.barcode.isnot(None), Product.barcode != "", func.lower(Product.barcode) == prod_key.lower()),
                )
            ).first()
            if not product and prod_key:
                product = db.query(Product).filter(
                    Product.name.isnot(None),
                    func.lower(func.trim(Product.name)) == prod_key.strip().lower(),
                ).first()
            if not warehouse:
                if wh_key and wh_key not in missing_warehouses:
                    missing_warehouses.append(wh_key)
                skipped += 1
                continue
            if not product:
                if prod_key and prod_key not in missing_products:
                    missing_products.append(prod_key)
                skipped += 1
                continue
            stock = db.query(Stock).filter(
                Stock.warehouse_id == warehouse.id,
                Stock.product_id == product.id,
            ).first()
            if stock:
                stock.quantity = qty
            else:
                db.add(Stock(warehouse_id=warehouse.id, product_id=product.id, quantity=qty))
            if tannarx is not None:
                product.purchase_price = tannarx
            if sotuv_narxi is not None:
                product.sale_price = sotuv_narxi
            updated += 1
        db.commit()
        detail = f"Yuklandi: {updated} ta"
        if skipped:
            detail += f", o'tkazib yuborildi: {skipped} ta"
            if missing_products:
                sample = ", ".join(missing_products[:5])
                if len(missing_products) > 5:
                    sample += f" va yana {len(missing_products) - 5} ta"
                detail += f". Mahsulot topilmadi: {sample}"
            if missing_warehouses:
                sample = ", ".join(missing_warehouses[:3])
                if len(missing_warehouses) > 3:
                    sample += f" va yana {len(missing_warehouses) - 3} ta"
                detail += f". Ombor topilmadi: {sample}"
        return RedirectResponse(url="/warehouse?success=import&detail=" + quote(detail), status_code=303)
    except Exception as e:
        traceback.print_exc()
        return RedirectResponse(url="/warehouse?error=import&detail=" + quote(str(e)[:200]), status_code=303)


@router.get("/transfers", response_class=HTMLResponse)
async def warehouse_transfers_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    transfers = db.query(WarehouseTransfer).order_by(WarehouseTransfer.date.desc()).limit(200).all()
    if getattr(current_user, "role", None) == "manager":
        wh_ids = [w.id for w in _warehouses_for_user(db, current_user)]
        if wh_ids:
            transfers = [t for t in transfers if (t.from_warehouse_id in wh_ids or t.to_warehouse_id in wh_ids)]
    error = request.query_params.get("error")
    return templates.TemplateResponse("warehouse/transfers_list.html", {
        "request": request,
        "current_user": current_user,
        "transfers": transfers,
        "page_title": "Ombordan omborga o'tkazish",
        "error_message": unquote(error) if error else None,
    })


@router.get("/transfers/new", response_class=HTMLResponse)
async def warehouse_transfer_new(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    warehouses = _warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    stocks = db.query(Stock).filter(Stock.quantity > 0).all()
    stock_by_warehouse_product = {}
    for s in stocks:
        wid, pid = str(s.warehouse_id), str(s.product_id)
        if wid not in stock_by_warehouse_product:
            stock_by_warehouse_product[wid] = {}
        stock_by_warehouse_product[wid][pid] = s.quantity
    products_list = [{"id": p.id, "name": (p.name or ""), "code": (p.code or "")} for p in products]
    return templates.TemplateResponse("warehouse/transfer_form.html", {
        "request": request,
        "current_user": current_user,
        "transfer": None,
        "warehouses": warehouses,
        "products": products,
        "products_list": products_list,
        "stock_by_warehouse_product": stock_by_warehouse_product,
        "now": datetime.now(),
        "page_title": "Ombordan omborga o'tkazish (yaratish)",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


@router.get("/transfers/{transfer_id}", response_class=HTMLResponse)
async def warehouse_transfer_edit(
    request: Request,
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    warehouses = _warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    stocks = db.query(Stock).filter(Stock.quantity > 0).all()
    stock_by_warehouse_product = {}
    for s in stocks:
        wid, pid = str(s.warehouse_id), str(s.product_id)
        if wid not in stock_by_warehouse_product:
            stock_by_warehouse_product[wid] = {}
        stock_by_warehouse_product[wid][pid] = s.quantity
    products_list = [{"id": p.id, "name": (p.name or ""), "code": (p.code or "")} for p in products]
    return templates.TemplateResponse("warehouse/transfer_form.html", {
        "request": request,
        "current_user": current_user,
        "transfer": transfer,
        "warehouses": warehouses,
        "products": products,
        "products_list": products_list,
        "stock_by_warehouse_product": stock_by_warehouse_product,
        "now": transfer.date or datetime.now(),
        "page_title": f"O'tkazish {transfer.number}",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


@router.post("/transfers/create")
async def warehouse_transfer_create(
    request: Request,
    from_warehouse_id: int = Form(...),
    to_warehouse_id: int = Form(...),
    note: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if from_warehouse_id == to_warehouse_id:
        return RedirectResponse(url="/warehouse/transfers/new?error=" + quote("Qayerdan va qayerga bir xil bo'lmasin."), status_code=303)
    form = await request.form()
    today = datetime.now()
    count = db.query(WarehouseTransfer).filter(
        WarehouseTransfer.date >= today.replace(hour=0, minute=0, second=0)
    ).count()
    number = f"OT-{today.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"
    transfer = WarehouseTransfer(
        number=number,
        from_warehouse_id=from_warehouse_id,
        to_warehouse_id=to_warehouse_id,
        status="draft",
        user_id=current_user.id,
        note=note or None,
    )
    db.add(transfer)
    db.commit()
    db.refresh(transfer)
    for key, value in form.items():
        if key.startswith("product_id_") and value:
            try:
                pid = int(value)
                qkey = "quantity_" + key.replace("product_id_", "")
                qty = float(form.get(qkey, "0").replace(",", "."))
                if pid and qty > 0:
                    db.add(WarehouseTransferItem(transfer_id=transfer.id, product_id=pid, quantity=qty))
            except (ValueError, TypeError):
                pass
    db.commit()
    return RedirectResponse(url=f"/warehouse/transfers/{transfer.id}", status_code=303)


@router.post("/transfers/{transfer_id}/save")
async def warehouse_transfer_save(
    request: Request,
    transfer_id: int,
    from_warehouse_id: int = Form(...),
    to_warehouse_id: int = Form(...),
    note: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer or transfer.status != "draft":
        raise HTTPException(status_code=404, detail="Hujjat topilmadi yoki tahrirlab bo'lmaydi")
    if from_warehouse_id == to_warehouse_id:
        return RedirectResponse(url=f"/warehouse/transfers/{transfer_id}?error=" + quote("Qayerdan va qayerga bir xil bo'lmasin."), status_code=303)
    transfer.from_warehouse_id = from_warehouse_id
    transfer.to_warehouse_id = to_warehouse_id
    transfer.note = note or None
    form = await request.form()
    db.query(WarehouseTransferItem).filter(WarehouseTransferItem.transfer_id == transfer_id).delete()
    for key, value in form.items():
        if key.startswith("product_id_") and value:
            try:
                pid = int(value)
                qkey = "quantity_" + key.replace("product_id_", "")
                qty = float(form.get(qkey, "0").replace(",", "."))
                if pid and qty > 0:
                    db.add(WarehouseTransferItem(transfer_id=transfer_id, product_id=pid, quantity=qty))
            except (ValueError, TypeError):
                pass
    db.commit()
    return RedirectResponse(url="/warehouse/transfers?saved=1", status_code=303)


@router.post("/transfers/{transfer_id}/confirm")
async def warehouse_transfer_confirm(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if transfer.status == "confirmed":
        return RedirectResponse(url=f"/warehouse/transfers/{transfer_id}?error=" + quote("Hujjat allaqachon tasdiqlangan."), status_code=303)
    items = db.query(WarehouseTransferItem).filter(WarehouseTransferItem.transfer_id == transfer_id).all()
    if not items:
        return RedirectResponse(url=f"/warehouse/transfers/{transfer_id}?error=" + quote("Kamida bitta mahsulot qo'shing."), status_code=303)
    for item in items:
        src = db.query(Stock).filter(
            Stock.warehouse_id == transfer.from_warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        need = float(item.quantity or 0)
        have = float(src.quantity or 0) if src else 0
        if not src or (have + 1e-6 < need):
            prod = db.query(Product).filter(Product.id == item.product_id).first()
            name = prod.name if prod else f"#{item.product_id}"
            avail_display = "0" if abs(have) < 1e-6 else ("%.6f" % have).rstrip("0").rstrip(".")
            return RedirectResponse(
                url=f"/warehouse/transfers/{transfer_id}?error=" + quote(f"Qayerdan omborda «{name}» yetarli emas (kerak: {item.quantity}, mavjud: {avail_display})"),
                status_code=303,
            )
    for item in items:
        src = db.query(Stock).filter(
            Stock.warehouse_id == transfer.from_warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        src.quantity -= item.quantity
        if src.quantity <= 0:
            src.quantity = 0
        dest = db.query(Stock).filter(
            Stock.warehouse_id == transfer.to_warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        if dest:
            dest.quantity += item.quantity
        else:
            db.add(Stock(warehouse_id=transfer.to_warehouse_id, product_id=item.product_id, quantity=item.quantity))
    transfer.status = "confirmed"
    db.commit()
    return RedirectResponse(url=f"/warehouse/transfers/{transfer_id}?confirmed=1", status_code=303)


@router.post("/transfers/{transfer_id}/revert")
async def warehouse_transfer_revert(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if transfer.status != "confirmed":
        return RedirectResponse(url="/warehouse/transfers?error=" + quote("Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin."), status_code=303)
    items = db.query(WarehouseTransferItem).filter(WarehouseTransferItem.transfer_id == transfer_id).all()
    for item in items:
        dest = db.query(Stock).filter(
            Stock.warehouse_id == transfer.to_warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        if dest:
            dest.quantity -= item.quantity
            if dest.quantity < 0:
                dest.quantity = 0
        src = db.query(Stock).filter(
            Stock.warehouse_id == transfer.from_warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        if src:
            src.quantity += item.quantity
        else:
            db.add(Stock(warehouse_id=transfer.from_warehouse_id, product_id=item.product_id, quantity=item.quantity))
    transfer.status = "draft"
    db.commit()
    return RedirectResponse(url="/warehouse/transfers?reverted=1", status_code=303)


@router.post("/transfers/{transfer_id}/delete")
async def warehouse_transfer_delete(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if transfer.status == "confirmed":
        return RedirectResponse(
            url="/warehouse/transfers?error=" + quote("Tasdiqlangan hujjatni to'g'ridan-to'g'ri o'chirib bo'lmaydi. Avval tasdiqni bekor qiling."),
            status_code=303,
        )
    db.delete(transfer)
    db.commit()
    return RedirectResponse(url="/warehouse/transfers?deleted=1", status_code=303)


@router.get("/movement", response_class=HTMLResponse)
async def warehouse_movement(request: Request, current_user: User = Depends(require_auth)):
    return RedirectResponse(url="/warehouse/transfers", status_code=302)


@router.post("/transfer")
async def warehouse_transfer(
    request: Request,
    from_warehouse_id: int = Form(...),
    to_warehouse_id: int = Form(...),
    product_id: int = Form(...),
    quantity: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if from_warehouse_id == to_warehouse_id:
        return RedirectResponse(url="/warehouse/movement?error=1&detail=" + quote("Qayerdan va qayerga ombor bir xil bo'lmasin."), status_code=303)
    if quantity <= 0:
        return RedirectResponse(url="/warehouse/movement?error=1&detail=" + quote("Miqdor 0 dan katta bo'lishi kerak."), status_code=303)
    source = db.query(Stock).filter(
        Stock.warehouse_id == from_warehouse_id,
        Stock.product_id == product_id,
    ).first()
    need_q = float(quantity or 0)
    have_q = float(source.quantity or 0) if source else 0
    if not source or (have_q + 1e-6 < need_q):
        product = db.query(Product).filter(Product.id == product_id).first()
        name = product.name if product else f"#{product_id}"
        avail_display = "0" if abs(have_q) < 1e-6 else ("%.6f" % have_q).rstrip("0").rstrip(".")
        return RedirectResponse(
            url="/warehouse/movement?error=1&detail=" + quote(f"Qayerdan omborda «{name}» yetarli emas (kerak: {quantity}, mavjud: {avail_display})"),
            status_code=303,
        )
    source.quantity -= quantity
    if source.quantity <= 0:
        source.quantity = 0
    dest = db.query(Stock).filter(
        Stock.warehouse_id == to_warehouse_id,
        Stock.product_id == product_id,
    ).first()
    if dest:
        dest.quantity += quantity
    else:
        db.add(Stock(warehouse_id=to_warehouse_id, product_id=product_id, quantity=quantity))
    db.commit()
    return RedirectResponse(url="/warehouse/movement?success=1", status_code=303)
