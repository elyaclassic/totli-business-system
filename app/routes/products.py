"""
Tovarlar — ro'yxat, qo'shish, tahrirlash, barcode, import/export.
"""
import io
import os
from typing import Optional
from urllib.parse import quote, unquote
from zipfile import BadZipFile

import openpyxl
import barcode
from barcode.writer import ImageWriter
from fastapi import APIRouter, Request, Depends, Form, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_

from app.core import templates
from app.models.database import get_db, Product, Category, Unit, User
from app.deps import require_auth, require_admin

router = APIRouter(prefix="/products", tags=["products"])


@router.get("/barcode/{product_id}")
async def product_barcode(product_id: int, download: int = 0, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product or not product.barcode:
        return HTMLResponse("<h3>Shtixkod topilmadi</h3>", status_code=404)
    barcode_path = f"app/static/images/products/barcode_{product.id}.png"
    if not os.path.exists(barcode_path):
        code128 = barcode.get("code128", product.barcode, writer=ImageWriter())
        code128.save(barcode_path[:-4])
    if download:
        return FileResponse(barcode_path, media_type="image/png", filename=f"barcode_{product.code}.png")
    return FileResponse(barcode_path, media_type="image/png")


@router.get("/export")
async def export_products(db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    products = db.query(Product).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["ID", "Kod", "Nomi", "Turi", "O'lchov", "Sotish narxi", "Olish narxi"])
    for p in products:
        ws.append([
            p.id, p.code, p.name, p.type,
            p.unit.name if p.unit else "",
            p.sale_price, p.purchase_price,
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products.xlsx"},
    )


@router.get("/template")
async def product_import_template(current_user: User = Depends(require_auth)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"
    headers = ["ID", "Kod", "Nomi", "Turi", "O'lchov", "Sotish narxi", "Olish narxi"]
    ws.append(headers)
    ws.append(["", "P001", "Tayyor mahsulot", "tayyor", "dona", 15000, 10000])
    ws.append(["", "P002", "Yarim tayyor mahsulot", "yarim_tayyor", "kg", 8000, 5000])
    ws.append(["", "P003", "Xom ashyo", "hom_ashyo", "kg", 2000, 1500])
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    ws.column_dimensions["C"].width = 30
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tovar_andoza.xlsx"},
    )


@router.get("/check", response_class=HTMLResponse)
async def products_check_duplicates(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovarlar tekshiruvi: bir xil nomdagi tovarlar (dublikatlar) ro'yxati."""
    by_name: dict = {}
    for p in db.query(Product).filter(Product.is_active == True, Product.name.isnot(None)).all():
        key = (p.name or "").strip().lower()
        if key:
            by_name.setdefault(key, []).append(p)
    duplicate_groups = [prods for prods in by_name.values() if len(prods) > 1]
    return templates.TemplateResponse("products/check.html", {
        "request": request,
        "current_user": current_user,
        "page_title": "Tovarlar tekshiruvi",
        "duplicate_groups": duplicate_groups,
    })


@router.get("/import")
async def products_import_get(current_user: User = Depends(require_auth)):
    return RedirectResponse(url="/products", status_code=303)


@router.post("/import")
async def import_products(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    form = await request.form()
    file = form.get("file") or form.get("excel_file")
    if not file or not getattr(file, "filename", None):
        return RedirectResponse(url="/products?error=import&detail=" + quote("Excel fayl tanlang"), status_code=303)
    try:
        contents = await file.read()
        if not contents:
            return RedirectResponse(url="/products?error=import&detail=" + quote("Fayl bo'sh"), status_code=303)
        if contents[:2] != b"PK":
            return RedirectResponse(
                url="/products?error=import&detail=" + quote("Fayl .xlsx formati bo'lishi kerak (Excel 2007+)."),
                status_code=303,
            )
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=False, data_only=True)
        ws = wb.active
        if ws.max_row < 2:
            return RedirectResponse(
                url="/products?import_ok=0&detail=" + quote("Excelda ma'lumot qatorlari yo'q."),
                status_code=303,
            )
        added, updated = 0, 0
        for row_num in range(2, ws.max_row + 1):
            def cell(col):
                v = ws.cell(row=row_num, column=col).value
                return "" if v is None else str(v).strip()

            code = cell(2) or cell(1)
            name = cell(3) or cell(2)
            if not code and not name:
                continue
            if code.lower() in ("id", "kod", "nomi", "turi", "o'lchov") and (not name or name.lower() in ("id", "kod", "nomi", "turi")):
                continue
            if not code:
                code = f"P{row_num}"
            if not name:
                name = code
            raw = (cell(4) or "tayyor").replace("\xa0", " ").strip().lower()
            if raw in ("yarim tayyor", "yarim_tayyor", "yarimtayyor"):
                type_ = "yarim_tayyor"
            elif raw in ("xom ashyo", "hom_ashyo", "xom_ashyo", "xomashyo"):
                type_ = "hom_ashyo"
            else:
                type_ = "tayyor"
            unit_name = cell(5) or None
            try:
                sale_price = float((cell(6) or "0").replace(" ", "").replace(",", "."))
            except (ValueError, TypeError):
                sale_price = 0
            try:
                purchase_price = float((cell(7) or "0").replace(" ", "").replace(",", "."))
            except (ValueError, TypeError):
                purchase_price = 0
            try:
                unit = db.query(Unit).filter(Unit.name == unit_name).first() if unit_name else None
                if not unit and unit_name:
                    unit = Unit(name=unit_name, code=unit_name.lower().replace(" ", "_")[:10] or "u")
                    db.add(unit)
                    db.commit()
                    db.refresh(unit)
                product = db.query(Product).filter(Product.code == code).first()
                if not product:
                    product = Product(code=code, is_active=True)
                    db.add(product)
                    added += 1
                else:
                    updated += 1
                product.name = name
                product.type = type_
                product.is_active = True
                product.category_id = None
                product.unit_id = unit.id if unit else None
                product.sale_price = sale_price
                product.purchase_price = purchase_price
                db.commit()
            except Exception:
                db.rollback()
                continue
        if added == 0 and updated == 0:
            return RedirectResponse(
                url="/products?import_ok=0&detail=" + quote("Hech qanday qator import qilinmadi."),
                status_code=303,
            )
        return RedirectResponse(url="/products?import_ok=1&added=" + str(added) + "&updated=" + str(updated), status_code=303)
    except BadZipFile:
        return RedirectResponse(
            url="/products?error=import&detail=" + quote("Fayl .xlsx formati bo'lishi kerak."),
            status_code=303,
        )
    except Exception as e:
        err_msg = str(e)[:200]
        if "zip" in err_msg.lower() or "not a zip" in err_msg.lower():
            err_msg = "Fayl .xlsx formati bo'lishi kerak."
        return RedirectResponse(url="/products?error=import&detail=" + quote(err_msg), status_code=303)


@router.get("", response_class=HTMLResponse)
async def products_list(
    request: Request,
    type: str = "all",
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    query = db.query(Product).options(joinedload(Product.unit)).filter(Product.is_active == True)
    if type == "tayyor":
        query = query.filter(Product.type == "tayyor")
    elif type == "yarim_tayyor":
        query = query.filter(Product.type == "yarim_tayyor")
    elif type == "hom_ashyo":
        query = query.filter(Product.type == "hom_ashyo")
    search_q = (q or "").strip()
    if search_q:
        like = f"%{search_q}%"
        query = query.filter(
            or_(
                Product.name.ilike(like),
                Product.code.ilike(like),
                Product.barcode.ilike(like),
            )
        )
    products = query.all()
    categories = db.query(Category).all()
    units = db.query(Unit).all()
    import_ok = request.query_params.get("import_ok")
    added = request.query_params.get("added")
    updated = request.query_params.get("updated")
    import_error = request.query_params.get("error") == "import"
    import_detail = unquote(request.query_params.get("detail", "") or "")
    return templates.TemplateResponse("products/list.html", {
        "request": request,
        "products": products,
        "categories": categories,
        "units": units,
        "current_type": type,
        "search_q": search_q,
        "current_user": current_user,
        "page_title": "Tovarlar",
        "import_ok": import_ok,
        "import_added": added,
        "import_updated": updated,
        "import_error": import_error,
        "import_detail": import_detail,
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) in ("admin", "rahbar", "raxbar"),
    })


@router.get("/{product_id}", response_class=HTMLResponse)
async def product_detail(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        return HTMLResponse("<h3>Mahsulot topilmadi</h3>", status_code=404)
    return templates.TemplateResponse("products/detail.html", {
        "request": request,
        "product": product,
        "current_user": current_user,
        "page_title": product.name or "Tovar",
    })


@router.post("/add")
async def product_add(
    request: Request,
    name: str = Form(...),
    type: str = Form(...),
    category_id: int = Form(None),
    unit_id: int = Form(None),
    barcode: str = Form(None),
    sale_price: float = Form(0),
    purchase_price: float = Form(0),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    import shutil
    product = Product(
        name=name,
        code=None,
        type=type,
        category_id=category_id if category_id and category_id > 0 else None,
        unit_id=unit_id if unit_id and unit_id > 0 else None,
        barcode=barcode,
        sale_price=sale_price,
        purchase_price=purchase_price,
        image=None,
    )
    db.add(product)
    db.commit()
    product.code = f"P{product.id:05d}"
    db.commit()
    if image and (image.filename or "").strip():
        ext = (image.filename or "").split(".")[-1] if "." in (image.filename or "") else "jpg"
        image_filename = f"{product.code}.{ext}"
        image_path = os.path.join("app", "static", "images", "products", image_filename)
        os.makedirs(os.path.dirname(image_path), exist_ok=True)
        with open(image_path, "wb") as buffer:
            shutil.copyfileobj(image.file, buffer)
        product.image = image_filename
        db.commit()
    return RedirectResponse(url="/products", status_code=303)


@router.post("/edit/{product_id}")
async def product_edit(
    product_id: int,
    name: str = Form(...),
    type: str = Form(...),
    category_id: int = Form(None),
    unit_id: int = Form(None),
    barcode: str = Form(None),
    sale_price: float = Form(0),
    purchase_price: float = Form(0),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    import shutil
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Mahsulot topilmadi")
    product.name = name
    product.type = type
    product.category_id = category_id if category_id and category_id > 0 else None
    product.unit_id = unit_id if unit_id and unit_id > 0 else None
    product.barcode = barcode or None
    product.sale_price = sale_price
    product.purchase_price = purchase_price
    if image and (image.filename or "").strip():
        ext = (image.filename or "").split(".")[-1] if "." in (image.filename or "") else "jpg"
        image_filename = f"{product.code}.{ext}"
        image_path = os.path.join("app", "static", "images", "products", image_filename)
        os.makedirs(os.path.dirname(image_path), exist_ok=True)
        with open(image_path, "wb") as buffer:
            shutil.copyfileobj(image.file, buffer)
        product.image = image_filename
    db.commit()
    return RedirectResponse(url="/products", status_code=303)


@router.post("/delete/{product_id}")
async def product_delete(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Mahsulot topilmadi")
    product.is_active = False
    db.commit()
    return RedirectResponse(url="/products", status_code=303)


@router.post("/delete-bulk")
async def product_delete_bulk(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Faqat admin: tanlangan tovarlarni o'chirish (is_active=False)."""
    form = await request.form()
    ids = form.getlist("product_ids")
    deleted = 0
    for sid in ids:
        try:
            pid = int(sid)
            product = db.query(Product).filter(Product.id == pid).first()
            if product:
                product.is_active = False
                deleted += 1
        except (ValueError, TypeError):
            pass
    db.commit()
    msg = quote(f"Tanlangan {deleted} ta tovar o'chirildi.") if deleted else quote("Hech narsa tanlanmadi.")
    return RedirectResponse(url="/products?deleted=" + msg, status_code=303)


@router.get("/bulk-update")
async def product_bulk_update_get(
    current_user: User = Depends(require_auth),
):
    """GET /products/bulk-update — brauzerda ochilganda tovarlar sahifasiga yo'naltirish."""
    return RedirectResponse(url="/products", status_code=303)


@router.post("/bulk-update")
async def product_bulk_update(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Faqat admin: tanlangan tovarlarga gruppaviy o'zgartirish (turi, o'lchov birligi)."""
    form = await request.form()
    ids = form.getlist("product_ids")
    new_type = (form.get("type") or "").strip()
    unit_id_raw = form.get("unit_id")
    new_unit_id = None
    if unit_id_raw is not None and str(unit_id_raw).strip() and str(unit_id_raw).strip() != "0":
        try:
            new_unit_id = int(unit_id_raw)
        except (ValueError, TypeError):
            pass
    updated = 0
    for sid in ids:
        try:
            pid = int(sid)
            product = db.query(Product).filter(Product.id == pid, Product.is_active == True).first()
            if not product:
                continue
            changed = False
            if new_type in ("tayyor", "yarim_tayyor", "hom_ashyo"):
                product.type = new_type
                changed = True
            if new_unit_id is not None:
                product.unit_id = new_unit_id
                changed = True
            if changed:
                updated += 1
        except (ValueError, TypeError):
            pass
    if updated > 0:
        db.commit()
    msg = quote(f"Tanlangan tovarlar yangilandi ({updated} ta).") if updated else quote("O'zgartirish kiritilmadi yoki tovarlar topilmadi.")
    return RedirectResponse(url="/products?updated=" + msg, status_code=303)


@router.post("/{product_id}/upload-image")
async def product_upload_image(
    product_id: int,
    image: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    import shutil
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Mahsulot topilmadi")
    if image and image.filename:
        ext = image.filename.split(".")[-1]
        image_filename = f"{product.code}.{ext}"
        image_path = os.path.join("app", "static", "images", "products", image_filename)
        os.makedirs(os.path.dirname(image_path), exist_ok=True)
        with open(image_path, "wb") as buffer:
            shutil.copyfileobj(image.file, buffer)
        product.image = image_filename
        db.commit()
    return RedirectResponse(url="/products", status_code=303)
