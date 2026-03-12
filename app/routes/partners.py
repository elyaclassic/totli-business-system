"""
Kontragentlar (partners) — ro'yxat, qo'shish, tahrir, o'chirish, export/import.
"""
import io
from fastapi import APIRouter, Request, Depends, Form, HTTPException, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
import openpyxl

from app.core import templates
from app.models.database import get_db, User, Partner, Order, Purchase
from app.deps import require_auth

router = APIRouter(prefix="/partners", tags=["partners"])


@router.get("", response_class=HTMLResponse)
async def partners_list(
    request: Request,
    type: str = "all",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    query = db.query(Partner)
    if type != "all":
        query = query.filter(Partner.type == type)
    partners = query.all()
    try:
        from app.config.maps_config import YANDEX_MAPS_API_KEY
        yandex_apikey = YANDEX_MAPS_API_KEY or ""
    except Exception:
        yandex_apikey = ""
    return templates.TemplateResponse("partners/list.html", {
        "request": request,
        "partners": partners,
        "current_type": type,
        "current_user": current_user,
        "page_title": "Kontragentlar",
        "yandex_maps_apikey": yandex_apikey,
    })


@router.post("/add")
async def partner_add(
    request: Request,
    name: str = Form(...),
    type: str = Form(...),
    phone: str = Form(""),
    address: str = Form(""),
    credit_limit: float = Form(0),
    discount_percent: float = Form(0),
    db: Session = Depends(get_db),
):
    existing_by_name = db.query(Partner).filter(Partner.name == name).first()
    if existing_by_name:
        raise HTTPException(status_code=400, detail=f"'{name}' nomli kontragent allaqachon mavjud!")
    if phone and phone.strip():
        existing_by_phone = db.query(Partner).filter(Partner.phone == phone).first()
        if existing_by_phone:
            raise HTTPException(status_code=400, detail=f"'{phone}' telefon raqamli kontragent allaqachon mavjud!")
    partner = Partner(
        name=name,
        code=None,
        type=type,
        phone=phone,
        address=address,
        credit_limit=credit_limit,
        discount_percent=discount_percent,
    )
    db.add(partner)
    db.commit()
    return RedirectResponse(url="/partners", status_code=303)


@router.post("/edit/{partner_id}")
async def partner_edit(
    partner_id: int,
    name: str = Form(...),
    type: str = Form(...),
    phone: str = Form(""),
    address: str = Form(""),
    credit_limit: float = Form(0),
    discount_percent: float = Form(0),
    db: Session = Depends(get_db),
):
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Kontragent topilmadi")
    existing_by_name = db.query(Partner).filter(Partner.name == name, Partner.id != partner_id).first()
    if existing_by_name:
        raise HTTPException(status_code=400, detail=f"'{name}' nomli kontragent allaqachon mavjud!")
    if phone and phone.strip():
        existing_by_phone = db.query(Partner).filter(Partner.phone == phone, Partner.id != partner_id).first()
        if existing_by_phone:
            raise HTTPException(status_code=400, detail=f"'{phone}' telefon raqamli kontragent allaqachon mavjud!")
    partner.name = name
    partner.type = type
    partner.phone = phone
    partner.address = address
    partner.credit_limit = credit_limit
    partner.discount_percent = discount_percent
    db.commit()
    return RedirectResponse(url="/partners", status_code=303)


@router.post("/delete/{partner_id}")
async def partner_delete(
    partner_id: int,
    db: Session = Depends(get_db),
):
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Kontragent topilmadi")
    has_orders = db.query(Order).filter(Order.partner_id == partner_id).first()
    has_purchases = db.query(Purchase).filter(Purchase.partner_id == partner_id).first()
    if has_orders or has_purchases:
        raise HTTPException(
            status_code=400,
            detail="Bu kontragent bilan bog'liq buyurtmalar yoki kirimlar mavjud. O'chirish mumkin emas.",
        )
    db.delete(partner)
    db.commit()
    return RedirectResponse(url="/partners", status_code=303)


@router.get("/export")
async def export_partners(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    partners = db.query(Partner).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partners"
    ws.append(["ID", "Kod", "Nomi", "Turi", "Telefon", "Manzil", "Kredit Limit", "Chegirma %"])
    for p in partners:
        ws.append([p.id, p.code, p.name, p.type, p.phone, p.address, p.credit_limit, p.discount_percent])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kontragentlar.xlsx"},
    )


@router.get("/template")
async def template_partners():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Nomi", "Turi", "Telefon", "Manzil", "Kredit Limit", "Chegirma %"])
    ws.append(["Mijoz MCHJ", "customer", "+998901234567", "Toshkent", 1000000, 0])
    ws.append(["Yetkazib Beruvchi", "supplier", "+998909876543", "Samarqand", 0, 0])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kontragent_andoza.xlsx"},
    )


@router.post("/import")
async def import_partners(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row[0]:
            continue
        name, type_, phone, address, credit_limit, discount_percent = (row[0:6] if len(row) >= 6 else row + [None] * (6 - len(row)))[:6]
        if name is None:
            continue
        partner = db.query(Partner).filter(Partner.name == name).first()
        if not partner:
            count = db.query(Partner).count()
            code = f"P{count + 1:04d}"
            partner = Partner(
                code=code,
                name=name,
                type=type_ or "customer",
                phone=phone or "",
                address=address or "",
                credit_limit=credit_limit or 0,
                discount_percent=discount_percent or 0,
            )
            db.add(partner)
        else:
            if phone is not None:
                partner.phone = phone
            if address is not None:
                partner.address = address
            if credit_limit is not None:
                partner.credit_limit = credit_limit
            if discount_percent is not None:
                partner.discount_percent = discount_percent
        db.commit()
    return RedirectResponse(url="/partners", status_code=303)
