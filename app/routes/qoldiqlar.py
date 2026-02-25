"""
Qoldiqlar — kassa, tovar, kontragent qoldiqlari va hujjatlar (1C uslubida).
"""
import io
import traceback
from datetime import datetime
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Request, Depends, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core import templates
from app.utils.user_scope import get_warehouses_for_user
from app.models.database import (
    get_db,
    User,
    Product,
    Partner,
    Warehouse,
    Stock,
    CashRegister,
    StockAdjustmentDoc,
    StockAdjustmentDocItem,
    CashBalanceDoc,
    CashBalanceDocItem,
    PartnerBalanceDoc,
    PartnerBalanceDocItem,
)
from app.deps import require_auth, require_admin

router = APIRouter(prefix="/qoldiqlar", tags=["qoldiqlar"])


@router.get("", response_class=HTMLResponse)
async def qoldiqlar_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Qoldiqlar sahifasi: kassa, tovar (forma spiska 1C), kontragent qoldiqlarini kiritish"""
    cash_registers = db.query(CashRegister).filter(CashRegister.is_active == True).all()
    warehouses = get_warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    stocks = db.query(Stock).join(Warehouse).join(Product).order_by(Stock.updated_at.desc()).limit(300).all()
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    tovar_docs = (
        db.query(StockAdjustmentDoc)
        .order_by(StockAdjustmentDoc.id.desc())
        .limit(500)
        .all()
    )
    cash_docs = (
        db.query(CashBalanceDoc)
        .order_by(CashBalanceDoc.created_at.desc())
        .limit(200)
        .all()
    )
    kontragent_docs = (
        db.query(PartnerBalanceDoc)
        .order_by(PartnerBalanceDoc.created_at.desc())
        .limit(200)
        .all()
    )
    return templates.TemplateResponse("qoldiqlar/index.html", {
        "request": request,
        "cash_registers": cash_registers,
        "warehouses": warehouses,
        "products": products,
        "stocks": stocks,
        "partners": partners,
        "tovar_docs": tovar_docs,
        "cash_docs": cash_docs,
        "kontragent_docs": kontragent_docs,
        "current_user": current_user,
        "page_title": "Qoldiqlar",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) in ("admin", "rahbar", "raxbar"),
    })


@router.post("/kassa/{cash_id}")
async def qoldiqlar_kassa_save(
    cash_id: int,
    balance: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kassa qoldig'ini yangilash (eski tezkor forma uchun qolgan)"""
    cash = db.query(CashRegister).filter(CashRegister.id == cash_id).first()
    if not cash:
        raise HTTPException(status_code=404, detail="Kassa topilmadi")
    cash.balance = balance
    db.commit()
    return RedirectResponse(url="/qoldiqlar#kassa", status_code=303)


# --- Kassa qoldiq HUJJATLARI (1C uslubida) ---
@router.get("/kassa/hujjat/new", response_class=HTMLResponse)
async def qoldiqlar_kassa_hujjat_new(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Yangi kassa qoldiq hujjati"""
    cash_registers = db.query(CashRegister).filter(CashRegister.is_active == True).all()
    return templates.TemplateResponse("qoldiqlar/kassa_hujjat_form.html", {
        "request": request,
        "doc": None,
        "cash_registers": cash_registers,
        "current_user": current_user,
        "page_title": "Kassa qoldiqlari — yangi hujjat",
    })


@router.post("/kassa/hujjat")
async def qoldiqlar_kassa_hujjat_create(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kassa qoldiq hujjatini yaratish (qoralama)"""
    form = await request.form()
    cash_ids = form.getlist("cash_register_id")
    balances = form.getlist("balance")

    items_data = []
    for i, cid in enumerate(cash_ids):
        if not cid:
            continue
        try:
            bid = int(cid)
            bal = float(balances[i]) if i < len(balances) and balances[i] != "" else None
        except (TypeError, ValueError):
            continue
        if bal is not None:
            items_data.append((bid, bal))

    if not items_data:
        return RedirectResponse(url="/qoldiqlar/kassa/hujjat/new", status_code=303)

    today = datetime.now()
    count = db.query(CashBalanceDoc).filter(
        CashBalanceDoc.date >= today.replace(hour=0, minute=0, second=0)
    ).count()
    number = f"KLD-{today.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"

    doc = CashBalanceDoc(
        number=number,
        date=today,
        user_id=current_user.id if current_user else None,
        status="draft",
    )
    db.add(doc)
    db.flush()
    for cid, bal in items_data:
        db.add(CashBalanceDocItem(doc_id=doc.id, cash_register_id=cid, balance=bal))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc.id}", status_code=303)


@router.get("/kassa/hujjat/{doc_id}", response_class=HTMLResponse)
async def qoldiqlar_kassa_hujjat_view(
    request: Request,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kassa qoldiq hujjatini ko'rish"""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    cash_registers = db.query(CashRegister).filter(CashRegister.is_active == True).all()
    return templates.TemplateResponse("qoldiqlar/kassa_hujjat_form.html", {
        "request": request,
        "doc": doc,
        "cash_registers": cash_registers,
        "current_user": current_user,
        "page_title": f"Kassa qoldiqlari {doc.number}",
    })


@router.post("/kassa/hujjat/{doc_id}/tasdiqlash")
async def qoldiqlar_kassa_hujjat_tasdiqlash(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kassa hujjatini tasdiqlash — kassa balanslarini yangilash"""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Hujjat allaqachon tasdiqlangan")
    if not doc.items:
        raise HTTPException(status_code=400, detail="Kamida bitta kassa qatori bo'lishi kerak")
    for item in doc.items:
        cash = db.query(CashRegister).filter(CashRegister.id == item.cash_register_id).first()
        if cash:
            item.previous_balance = cash.balance
            cash.balance = item.balance
    doc.status = "confirmed"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc_id}", status_code=303)


@router.post("/kassa/hujjat/{doc_id}/revert")
async def qoldiqlar_kassa_hujjat_revert(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kassa hujjati tasdiqini bekor qilish (faqat admin)"""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "confirmed":
        raise HTTPException(status_code=400, detail="Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin")
    for item in doc.items:
        cash = db.query(CashRegister).filter(CashRegister.id == item.cash_register_id).first()
        if cash and item.previous_balance is not None:
            cash.balance = item.previous_balance
    doc.status = "draft"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc_id}", status_code=303)


@router.post("/kassa/hujjat/{doc_id}/delete")
async def qoldiqlar_kassa_hujjat_delete(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kassa hujjatini o'chirish (faqat qoralama, faqat admin)"""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama holatidagi hujjatni o'chirish mumkin. Avval tasdiqni bekor qiling.")
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#kassa", status_code=303)


# --- Kontragent qoldiq HUJJATLARI (1C uslubida) ---
@router.get("/kontragent/hujjat/new", response_class=HTMLResponse)
async def qoldiqlar_kontragent_hujjat_new(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Yangi kontragent balans hujjati"""
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    return templates.TemplateResponse("qoldiqlar/kontragent_hujjat_form.html", {
        "request": request,
        "doc": None,
        "partners": partners,
        "current_user": current_user,
        "page_title": "Kontragent qoldiqlari — yangi hujjat",
    })


@router.post("/kontragent/hujjat")
async def qoldiqlar_kontragent_hujjat_create(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent balans hujjatini yaratish (qoralama)"""
    form = await request.form()
    partner_ids = form.getlist("partner_id")
    balances = form.getlist("balance")

    items_data = []
    for i, pid in enumerate(partner_ids):
        if not pid:
            continue
        try:
            pid_int = int(pid)
            bal_str = (balances[i] if i < len(balances) else "").strip()
            if not bal_str:
                continue
            bal = float(bal_str)
        except (TypeError, ValueError):
            continue
        items_data.append((pid_int, bal))

    if not items_data:
        return RedirectResponse(url="/qoldiqlar/kontragent/hujjat/new", status_code=303)

    today = datetime.now()
    count = db.query(PartnerBalanceDoc).filter(
        PartnerBalanceDoc.date >= today.replace(hour=0, minute=0, second=0)
    ).count()
    number = f"KNT-{today.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"

    doc = PartnerBalanceDoc(
        number=number,
        date=today,
        user_id=current_user.id if current_user else None,
        status="draft",
    )
    db.add(doc)
    db.flush()
    for pid, bal in items_data:
        db.add(PartnerBalanceDocItem(doc_id=doc.id, partner_id=pid, balance=bal))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc.id}", status_code=303)


@router.get("/kontragent/hujjat/{doc_id}", response_class=HTMLResponse)
async def qoldiqlar_kontragent_hujjat_view(
    request: Request,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent balans hujjatini ko'rish"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    return templates.TemplateResponse("qoldiqlar/kontragent_hujjat_form.html", {
        "request": request,
        "doc": doc,
        "partners": partners,
        "current_user": current_user,
        "page_title": f"Kontragent qoldiqlari {doc.number}",
    })


@router.post("/kontragent/hujjat/{doc_id}/tasdiqlash")
async def qoldiqlar_kontragent_hujjat_tasdiqlash(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent hujjatini tasdiqlash — kontragent balanslarini yangilash"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Hujjat allaqachon tasdiqlangan")
    if not doc.items:
        raise HTTPException(status_code=400, detail="Kamida bitta kontragent qatori bo'lishi kerak")
    for item in doc.items:
        partner = db.query(Partner).filter(Partner.id == item.partner_id).first()
        if partner:
            item.previous_balance = partner.balance
            partner.balance = item.balance
    doc.status = "confirmed"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}", status_code=303)


@router.post("/kontragent/hujjat/{doc_id}/revert")
async def qoldiqlar_kontragent_hujjat_revert(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kontragent hujjati tasdiqini bekor qilish (faqat admin)"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "confirmed":
        raise HTTPException(status_code=400, detail="Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin")
    for item in doc.items:
        partner = db.query(Partner).filter(Partner.id == item.partner_id).first()
        if partner and item.previous_balance is not None:
            partner.balance = item.previous_balance
    doc.status = "draft"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}", status_code=303)


@router.post("/kontragent/hujjat/{doc_id}/delete")
async def qoldiqlar_kontragent_hujjat_delete(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kontragent hujjatini o'chirish (faqat qoralama, faqat admin)"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama holatidagi hujjatni o'chirish mumkin. Avval tasdiqni bekor qiling.")
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#kontragent", status_code=303)


@router.post("/tovar")
async def qoldiqlar_tovar_save(
    warehouse_id: int = Form(...),
    product_id: int = Form(...),
    quantity: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldig'ini kiritish yoki qo'shish (omborda mavjud bo'lsa qo'shiladi)"""
    if quantity < 0:
        return RedirectResponse(url="/qoldiqlar#tovar", status_code=303)
    stock = db.query(Stock).filter(
        Stock.warehouse_id == warehouse_id,
        Stock.product_id == product_id,
    ).first()
    if stock:
        stock.quantity = (stock.quantity or 0) + quantity
        stock.updated_at = datetime.now()
    else:
        stock = Stock(warehouse_id=warehouse_id, product_id=product_id, quantity=quantity)
        db.add(stock)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#tovar", status_code=303)


@router.post("/kontragent/{partner_id}")
async def qoldiqlar_kontragent_save(
    partner_id: int,
    balance: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent balansini yangilash"""
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Kontragent topilmadi")
    balance_str = (balance or "").strip()
    if not balance_str:
        return RedirectResponse(url="/qoldiqlar#kontragent", status_code=303)
    try:
        partner.balance = float(balance_str)
    except (TypeError, ValueError):
        return RedirectResponse(url="/qoldiqlar#kontragent", status_code=303)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#kontragent", status_code=303)


@router.get("/export")
async def qoldiqlar_export(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldiqlari hisoboti — Excel hujjat sifatida yuklab olish"""
    stocks = (
        db.query(Stock)
        .join(Warehouse, Stock.warehouse_id == Warehouse.id)
        .join(Product, Stock.product_id == Product.id)
        .order_by(Warehouse.name, Product.name)
        .all()
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tovar qoldiqlari"
    ws.append(["Ombor", "Mahsulot", "Kod", "Miqdor"])
    for s in stocks:
        ws.append([
            s.warehouse.name if s.warehouse else "-",
            s.product.name if s.product else "-",
            (s.product.code or "") if s.product else "",
            float(s.quantity) if s.quantity is not None else 0,
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tovar_qoldiqlari.xlsx"},
    )


# --- Tovar qoldiq HUJJATLARI (1C uslubida: ro'yxat + hujjat + qatorlar) ---
@router.get("/tovar/hujjat", response_class=HTMLResponse)
async def qoldiqlar_tovar_hujjat_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldiqlari hujjatlari ro'yxati"""
    docs = (
        db.query(StockAdjustmentDoc)
        .order_by(StockAdjustmentDoc.created_at.desc())
        .limit(200)
        .all()
    )
    return templates.TemplateResponse("qoldiqlar/hujjat_list.html", {
        "request": request,
        "docs": docs,
        "current_user": current_user,
        "page_title": "Tovar qoldiqlari hujjatlari",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) in ("admin", "rahbar", "raxbar"),
    })


@router.get("/tovar/hujjat/new", response_class=HTMLResponse)
async def qoldiqlar_tovar_hujjat_new(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Yangi tovar qoldiq hujjati (qoralama)"""
    warehouses = get_warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    return templates.TemplateResponse("qoldiqlar/hujjat_form.html", {
        "request": request,
        "doc": None,
        "warehouses": warehouses,
        "products": products,
        "current_user": current_user,
        "page_title": "Tovar qoldiqlari — yangi hujjat",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) in ("admin", "rahbar", "raxbar"),
    })


@router.post("/tovar/hujjat")
async def qoldiqlar_tovar_hujjat_create(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldiq hujjatini yaratish (qoralama)"""
    form = await request.form()
    product_ids = form.getlist("product_id")
    warehouse_ids = form.getlist("warehouse_id")
    quantities = form.getlist("quantity")
    cost_prices = form.getlist("cost_price")
    sale_prices = form.getlist("sale_price")

    items_data = []
    for i, pid in enumerate(product_ids):
        if not pid or not str(pid).strip():
            continue
        try:
            wid = int(warehouse_ids[i]) if i < len(warehouse_ids) and warehouse_ids[i] else None
            qty = float(quantities[i]) if i < len(quantities) and str(quantities[i]).strip() else 0
            _cp = cost_prices[i] if i < len(cost_prices) else ""
            _sp = sale_prices[i] if i < len(sale_prices) else ""
            cp = float(_cp) if str(_cp).strip() else 0
            sp = float(_sp) if str(_sp).strip() else 0
        except (TypeError, ValueError):
            continue
        if wid and qty > 0:
            try:
                items_data.append((int(pid), wid, qty, cp, sp))
            except ValueError:
                continue

    today = datetime.now()
    count = db.query(StockAdjustmentDoc).filter(
        StockAdjustmentDoc.date >= today.replace(hour=0, minute=0, second=0)
    ).count()
    number = f"QLD-{today.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"

    total_tannarx = sum(qty * cp for _, _, qty, cp, _ in items_data)
    total_sotuv = sum(qty * sp for _, _, qty, _, sp in items_data)

    doc = StockAdjustmentDoc(
        number=number,
        date=today,
        user_id=current_user.id if current_user else None,
        status="draft",
        total_tannarx=total_tannarx,
        total_sotuv=total_sotuv,
    )
    db.add(doc)
    db.flush()

    for pid, wid, qty, cp, sp in items_data:
        db.add(StockAdjustmentDocItem(
            doc_id=doc.id,
            product_id=pid,
            warehouse_id=wid,
            quantity=qty,
            cost_price=cp,
            sale_price=sp,
        ))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc.id}", status_code=303)


@router.post("/tovar/import-excel")
async def qoldiqlar_tovar_import_excel(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Exceldan tovar qoldiqlarini yuklash — hujjat (QLD-...) yaratiladi, jadvalda ko'rinadi."""
    form = await request.form()
    file = form.get("file") or form.get("excel_file")
    if not file or not getattr(file, "filename", None):
        return RedirectResponse(url="/qoldiqlar?error=import&detail=" + quote("Excel fayl tanlang") + "#tovar", status_code=303)
    try:
        contents = await file.read()
        if not contents:
            return RedirectResponse(url="/qoldiqlar?error=import&detail=" + quote("Fayl bo'sh") + "#tovar", status_code=303)
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=False, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        items_data = []
        for row in rows:
            if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                continue
            wh_key = str(row[0] or "").strip() if len(row) > 0 else ""
            raw_prod = row[1] if len(row) > 1 else None
            if raw_prod is not None and isinstance(raw_prod, (int, float)) and float(raw_prod) == int(float(raw_prod)):
                prod_key = str(int(float(raw_prod)))
            else:
                prod_key = str(raw_prod or "").strip()
            try:
                qty = float(row[2]) if len(row) > 2 and row[2] is not None else 0
            except (TypeError, ValueError):
                qty = 0
            cp = 0.0
            sp = 0.0
            if len(row) > 3 and row[3] is not None and row[3] != "":
                try:
                    cp = float(row[3])
                except (TypeError, ValueError):
                    pass
            if len(row) > 4 and row[4] is not None and row[4] != "":
                try:
                    sp = float(row[4])
                except (TypeError, ValueError):
                    pass
            if not wh_key or not prod_key or qty <= 0:
                continue
            warehouse = db.query(Warehouse).filter(
                (func.lower(Warehouse.name) == wh_key.lower()) | (Warehouse.code == wh_key)
            ).first()
            product = db.query(Product).filter(
                (Product.code == prod_key) | (Product.barcode == prod_key)
            ).first()
            if not product and prod_key:
                product = db.query(Product).filter(
                    Product.name.isnot(None),
                    func.lower(Product.name) == prod_key.lower()
                ).first()
            if not warehouse or not product:
                continue
            items_data.append((product.id, warehouse.id, qty, cp, sp))
        if not items_data:
            return RedirectResponse(
                url="/qoldiqlar?error=import&detail=" + quote("Hech qanday to'g'ri qator topilmadi. Ombor va mahsulot nomi/kodi to'g'ri ekanligini tekshiring.") + "#tovar",
                status_code=303,
            )
        today = datetime.now()
        count = db.query(StockAdjustmentDoc).filter(
            StockAdjustmentDoc.date >= today.replace(hour=0, minute=0, second=0)
        ).count()
        number = f"QLD-{today.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"
        total_tannarx = sum(qty * cp for _, _, qty, cp, _ in items_data)
        total_sotuv = sum(qty * sp for _, _, qty, _, sp in items_data)
        doc = StockAdjustmentDoc(
            number=number,
            date=today,
            user_id=current_user.id if current_user else None,
            status="draft",
            total_tannarx=total_tannarx,
            total_sotuv=total_sotuv,
        )
        db.add(doc)
        db.flush()
        for pid, wid, qty, cp, sp in items_data:
            db.add(StockAdjustmentDocItem(
                doc_id=doc.id,
                product_id=pid,
                warehouse_id=wid,
                quantity=qty,
                cost_price=cp,
                sale_price=sp,
            ))
        db.commit()
        return RedirectResponse(
            url="/qoldiqlar?success=import&doc_number=" + quote(doc.number) + "#tovar",
            status_code=303,
        )
    except Exception as e:
        traceback.print_exc()
        return RedirectResponse(
            url="/qoldiqlar?error=import&detail=" + quote(str(e)[:180]) + "#tovar",
            status_code=303,
        )


@router.get("/tovar/hujjat/{doc_id}", response_class=HTMLResponse)
async def qoldiqlar_tovar_hujjat_view(
    request: Request,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldiq hujjatini ko'rish/tahrirlash"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    warehouses = get_warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    return templates.TemplateResponse("qoldiqlar/hujjat_form.html", {
        "request": request,
        "doc": doc,
        "warehouses": warehouses,
        "products": products,
        "current_user": current_user,
        "page_title": f"Tovar qoldiqlari {doc.number}",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) in ("admin", "rahbar", "raxbar"),
    })


@router.post("/tovar/hujjat/{doc_id}/add-row")
async def qoldiqlar_tovar_hujjat_add_row(
    doc_id: int,
    product_id: int = Form(...),
    warehouse_id: int = Form(...),
    quantity: float = Form(...),
    cost_price: float = Form(0),
    sale_price: float = Form(0),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Hujjatga qator qo'shish (faqat qoralama)"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralamani tahrirlash mumkin")
    if quantity <= 0:
        return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)
    doc.total_tannarx = (doc.total_tannarx or 0) + quantity * (cost_price or 0)
    doc.total_sotuv = (doc.total_sotuv or 0) + quantity * (sale_price or 0)
    db.add(StockAdjustmentDocItem(
        doc_id=doc_id,
        product_id=product_id,
        warehouse_id=warehouse_id,
        quantity=quantity,
        cost_price=cost_price or 0,
        sale_price=sale_price or 0,
    ))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/delete-row/{item_id}")
async def qoldiqlar_tovar_hujjat_delete_row(
    doc_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Hujjatdan qatorni o'chirish (faqat qoralama)"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralamani tahrirlash mumkin")
    item = db.query(StockAdjustmentDocItem).filter(
        StockAdjustmentDocItem.id == item_id,
        StockAdjustmentDocItem.doc_id == doc_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Qator topilmadi")
    doc.total_tannarx = (doc.total_tannarx or 0) - (item.quantity * (item.cost_price or 0))
    doc.total_sotuv = (doc.total_sotuv or 0) - (item.quantity * (item.sale_price or 0))
    if doc.total_tannarx < 0:
        doc.total_tannarx = 0
    if doc.total_sotuv < 0:
        doc.total_sotuv = 0
    db.delete(item)
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/tasdiqlash")
async def qoldiqlar_tovar_hujjat_tasdiqlash(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Hujjatni tasdiqlash — ombor qoldig'i hujjatdagi miqdorga o'rnatiladi (qo'shilmaydi). Bir ombor+mahsulot uchun bitta Stock qatori qoladi."""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Hujjat allaqachon tasdiqlangan")
    if not doc.items:
        raise HTTPException(status_code=400, detail="Kamida bitta qator bo'lishi kerak")

    for item in doc.items:
        stocks = db.query(Stock).filter(
            Stock.warehouse_id == item.warehouse_id,
            Stock.product_id == item.product_id,
        ).all()
        new_quantity = item.quantity
        if stocks:
            keep = stocks[0]
            keep.quantity = new_quantity
            keep.updated_at = datetime.now()
            for s in stocks[1:]:
                db.delete(s)
        else:
            db.add(Stock(
                warehouse_id=item.warehouse_id,
                product_id=item.product_id,
                quantity=new_quantity,
            ))
    doc.status = "confirmed"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/revert")
async def qoldiqlar_tovar_hujjat_revert(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Tovar qoldiq hujjati tasdiqini bekor qilish (faqat admin) — ombor qoldig'ini kamaytirish"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "confirmed":
        raise HTTPException(status_code=400, detail="Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin")
    for item in doc.items:
        stock = db.query(Stock).filter(
            Stock.warehouse_id == item.warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        if stock:
            stock.quantity = (stock.quantity or 0) - item.quantity
            if stock.quantity < 0:
                stock.quantity = 0
            stock.updated_at = datetime.now()
    doc.status = "draft"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/delete")
async def qoldiqlar_tovar_hujjat_delete(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Tovar qoldiq hujjatini o'chirish (faqat qoralama, faqat admin)"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama holatidagi hujjatni o'chirish mumkin. Avval tasdiqni bekor qiling.")
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#tovar", status_code=303)
