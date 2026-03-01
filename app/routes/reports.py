"""
Hisobotlar — savdo, qoldiq, qarzdorlik va Excel export.
"""
import io
import json
from datetime import datetime, timedelta
from fastapi import APIRouter, Request, Depends, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core import templates
from app.models.database import get_db, Order, OrderItem, Stock, StockMovement, Product, Partner, Warehouse, User, Production, Recipe, StockAdjustmentDoc, StockAdjustmentDocItem, Employee, Purchase, PurchaseItem, WarehouseTransfer, Payment
from app.deps import get_current_user, require_auth, require_admin
from app.utils.user_scope import get_warehouses_for_user

router = APIRouter(prefix="/reports", tags=["reports"])


def get_allowed_report_types(user: User) -> list:
    """Foydalanuvchiga ruxsat berilgan hisobot turlarini qaytaradi."""
    if not user:
        return []
    # Admin uchun barcha hisobotlar
    if user.role == "admin":
        return ["sales", "stock", "debts", "production", "employees", "profit", "partner_reconciliation"]
    # allowed_sections bo'sh yoki None bo'lsa, hech narsa ko'rsatilmaydi
    if not user.allowed_sections:
        return []
    try:
        sections = json.loads(user.allowed_sections) if isinstance(user.allowed_sections, str) else user.allowed_sections
        if not isinstance(sections, list):
            return []
        # allowed_sections ichida "reports_sales", "reports_stock" kabi formatda bo'lishi mumkin
        report_types = []
        for s in sections:
            if isinstance(s, str) and s.startswith("reports_"):
                report_type = s.replace("reports_", "")
                if report_type in ["sales", "stock", "debts", "production", "employees", "profit", "partner_reconciliation"]:
                    report_types.append(report_type)
        return report_types
    except (json.JSONDecodeError, TypeError, AttributeError):
        return []


@router.get("", response_class=HTMLResponse)
async def reports_index(request: Request, current_user: User = Depends(require_auth)):
    """Hisobotlar bosh sahifasi"""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    allowed_types = get_allowed_report_types(current_user)
    return templates.TemplateResponse("reports/index.html", {
        "request": request,
        "page_title": "Hisobotlar",
        "current_user": current_user,
        "allowed_report_types": allowed_types,
    })


@router.get("/form", response_class=HTMLResponse)
async def reports_form(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    """Hisobotlar formasi — hisobot turi va filtrlarni tanlash"""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    today = datetime.now()
    start_date = today.replace(day=1).strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")
    warehouses = get_warehouses_for_user(db, current_user)
    allowed_types = get_allowed_report_types(current_user)
    return templates.TemplateResponse("reports/form.html", {
        "request": request,
        "page_title": "Hisobotlar formasi",
        "current_user": current_user,
        "start_date": start_date,
        "end_date": end_date,
        "warehouses": warehouses,
        "allowed_report_types": allowed_types,
    })


@router.get("/sales", response_class=HTMLResponse)
async def report_sales(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    orders = db.query(Order).filter(
        Order.type == "sale",
        Order.date >= start_date,
        Order.date <= end_date + " 23:59:59",
    ).all()
    total = sum(o.total for o in orders)
    return templates.TemplateResponse("reports/sales.html", {
        "request": request,
        "orders": orders,
        "total": total,
        "start_date": start_date,
        "end_date": end_date,
        "page_title": "Savdo hisoboti",
        "current_user": current_user,
    })


@router.get("/sales/export")
async def report_sales_export(
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    orders = db.query(Order).filter(
        Order.type == "sale",
        Order.date >= start_date,
        Order.date <= end_date + " 23:59:59",
    ).order_by(Order.date.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Savdo"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A1"] = "Savdo hisoboti"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Davr: {start_date} — {end_date}"
    ws.append(["№", "Sana", "Buyurtma", "Mijoz", "Jami", "Holat"])
    for c in range(1, 7):
        ws.cell(row=4, column=c).fill = header_fill
        ws.cell(row=4, column=c).font = Font(bold=True, color="FFFFFF")
    for i, o in enumerate(orders, 1):
        ws.append([
            i,
            o.date.strftime("%d.%m.%Y %H:%M") if o.date else "",
            o.number or "",
            o.partner.name if o.partner else "",
            float(o.total or 0),
            o.status or "",
        ])
    total = sum(o.total or 0 for o in orders)
    ws.append(["", "", "", "JAMI:", total, ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=savdo_{start_date}_{end_date}.xlsx"},
    )


@router.get("/stock", response_class=HTMLResponse)
async def report_stock(
    request: Request,
    warehouse_id: str = None,
    merged: int = None,
    cleared: int = None,
    recalculated: int = None,
    msg: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Qoldiq hisoboti — ombor bo'yicha qoldiqlar. Bir ombor + mahsulot uchun bitta qator (dublikatlar yig'indisi)."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    warehouses = get_warehouses_for_user(db, current_user)
    wh_ids = [w.id for w in warehouses]
    wh_id = None
    if warehouse_id is not None and str(warehouse_id).strip() != "":
        try:
            wid = int(warehouse_id)
            if not wh_ids or wid in wh_ids:
                wh_id = wid
        except (ValueError, TypeError):
            pass
    values = _stock_report_filtered(db, wh_id)
    stocks = [{"warehouse": v["warehouse"], "product": v["product"], "quantity": v["quantity"]} for v in values]
    # Jami summa: har bir qator uchun quantity * purchase_price yig'indisi (anniq)
    total_sum = 0.0
    for v in stocks:
        qty = float(v.get("quantity") or 0)
        price = float(getattr(v.get("product"), "purchase_price", None) or 0)
        total_sum += qty * price
    return templates.TemplateResponse("reports/stock.html", {
        "request": request,
        "stocks": stocks,
        "total_sum": total_sum,
        "warehouses": warehouses,
        "selected_warehouse_id": wh_id,
        "merged": merged,
        "cleared": cleared,
        "recalculated": recalculated,
        "msg": msg,
        "today": datetime.now().strftime("%Y-%m-%d"),
        "page_title": "Qoldiq hisoboti",
        "current_user": current_user,
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


def _document_type_label(doc_type: str) -> str:
    """Hujjat turi uchun o'qiladigan nom"""
    labels = {
        "Purchase": "Kirim (sotib olish)",
        "Production": "Ishlab chiqarish",
        "WarehouseTransfer": "Ombordan omborga",
        "StockAdjustmentDoc": "Qoldiq tuzatish",
        "Sale": "Sotuv",
        "SaleReturn": "Qaytish",
    }
    return labels.get(doc_type, doc_type or "—")


def _document_url(doc_type: str, doc_id: int) -> str:
    """Hujjat turi va ID bo'yicha ko'rish havolasi"""
    if doc_type == "Purchase":
        return f"/purchases/edit/{doc_id}"
    if doc_type == "Production":
        return f"/production/{doc_id}/materials"
    if doc_type == "WarehouseTransfer":
        return f"/warehouse/transfers/{doc_id}"
    if doc_type == "StockAdjustmentDoc":
        return f"/qoldiqlar/tovar/hujjat/{doc_id}"
    if doc_type == "Sale":
        return f"/sales/edit/{doc_id}"
    return "#"


@router.get("/stock/source", response_class=HTMLResponse)
async def report_stock_source(
    request: Request,
    warehouse_id: int = None,
    product_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Berilgan ombor + mahsulot uchun qoldiq manbai — barcha harakatlar (qaysi hujjatdan kirgan/chiqqan)."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not warehouse_id or not product_id:
        return RedirectResponse(url="/reports/stock", status_code=303)
    warehouse = db.query(Warehouse).filter(Warehouse.id == warehouse_id).first()
    product = db.query(Product).filter(Product.id == product_id).first()
    if not warehouse or not product:
        return RedirectResponse(url="/reports/stock", status_code=303)
    movements = (
        db.query(StockMovement)
        .filter(
            StockMovement.warehouse_id == warehouse_id,
            StockMovement.product_id == product_id,
        )
        .order_by(StockMovement.created_at.desc())
        .all()
    )
    # Faqat tasdiqlangan qoldiq tuzatish hujjatlarini ko'rsatamiz; qoralama/o'chirilganlarni olib tashlaymiz
    doc_ids = [m.document_id for m in movements if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id]
    doc_dates = {}
    confirmed_adj_ids = set()
    if doc_ids:
        for doc in db.query(StockAdjustmentDoc).filter(
            StockAdjustmentDoc.id.in_(doc_ids),
            StockAdjustmentDoc.status == "confirmed",
        ).all():
            doc_dates[doc.id] = doc.date
            confirmed_adj_ids.add(doc.id)
    # Bir xil hujjat (document_type, document_id) uchun bitta qator — dublikat harakatlar birlashtiriladi
    rows = []
    seen_doc = set()  # (document_type, document_id)
    for m in movements:
        # Qoldiq tuzatish: faqat tasdiqlangan hujjat ko'rinsin
        if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id not in confirmed_adj_ids:
            continue
        key = (m.document_type or "", m.document_id)
        if key in seen_doc:
            continue
        seen_doc.add(key)
        if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id and m.document_id in doc_dates and doc_dates[m.document_id]:
            display_date = doc_dates[m.document_id].strftime("%d.%m.%Y %H:%M")
        else:
            display_date = m.created_at.strftime("%d.%m.%Y %H:%M") if m.created_at else "—"
        rows.append({
            "date": display_date,
            "document_type": m.document_type or "",
            "document_type_label": _document_type_label(m.document_type or ""),
            "document_number": m.document_number or f"{m.document_type}-{m.document_id}",
            "document_id": m.document_id,
            "document_url": _document_url(m.document_type or "", m.document_id),
            "quantity_change": float(m.quantity_change or 0),
            "quantity_after": float(m.quantity_after or 0),
        })
    # Barcha hujjat turlarida sana — hujjat sanasi (harakat yozilgan vaqt emas), barcha mahsulotlar uchun bir xil
    purchase_ids = [r["document_id"] for r in rows if (r.get("document_type") or "") == "Purchase" and r.get("document_id")]
    sale_ids = [r["document_id"] for r in rows if (r.get("document_type") or "") in ("Sale", "SaleReturn", "SaleReturnRevert") and r.get("document_id")]
    transfer_ids = [r["document_id"] for r in rows if (r.get("document_type") or "") == "WarehouseTransfer" and r.get("document_id")]
    purchases_by_id = {p.id: p for p in db.query(Purchase).filter(Purchase.id.in_(purchase_ids)).all()} if purchase_ids else {}
    orders_by_id = {o.id: o for o in db.query(Order).filter(Order.id.in_(sale_ids)).all()} if sale_ids else {}
    transfers_by_id = {t.id: t for t in db.query(WarehouseTransfer).filter(WarehouseTransfer.id.in_(transfer_ids)).all()} if transfer_ids else {}
    for r in rows:
        doc_type = r.get("document_type") or ""
        doc_id = r.get("document_id")
        if doc_type == "Purchase" and doc_id and doc_id in purchases_by_id and purchases_by_id[doc_id].date:
            r["date"] = purchases_by_id[doc_id].date.strftime("%d.%m.%Y %H:%M")
        elif doc_type in ("Sale", "SaleReturn", "SaleReturnRevert") and doc_id and doc_id in orders_by_id and orders_by_id[doc_id].date:
            r["date"] = orders_by_id[doc_id].date.strftime("%d.%m.%Y %H:%M")
        elif doc_type == "WarehouseTransfer" and doc_id and doc_id in transfers_by_id and transfers_by_id[doc_id].date:
            r["date"] = transfers_by_id[doc_id].date.strftime("%d.%m.%Y %H:%M")
    # Ishlab chiqarish qatorlari uchun: hujjat sanasi va hujjatda ko'rsatilgan miqdor bilan harakatdagi miqdorni solishtirish (farq bo'lsa ogohlantirish)
    from sqlalchemy.orm import joinedload
    from app.utils.production_order import production_output_quantity_for_stock
    prod_ids = [r["document_id"] for r in rows if (r.get("document_type") or "") == "Production" and r.get("document_id")]
    productions_by_id = {}
    if prod_ids:
        for p in db.query(Production).options(joinedload(Production.recipe)).filter(Production.id.in_(prod_ids)).all():
            productions_by_id[p.id] = p
    for r in rows:
        if (r.get("document_type") or "") != "Production" or not r.get("document_id"):
            continue
        prod = productions_by_id.get(r["document_id"])
        if not prod:
            continue
        # Qoldiq manbai hisobotida sana — hujjat sanasi (Production.date), tasdiqlash vaqtida emas
        if prod.date:
            r["date"] = prod.date.strftime("%d.%m.%Y %H:%M")
        if not prod.recipe:
            continue
        expected = production_output_quantity_for_stock(db, prod, prod.recipe)
        change = r.get("quantity_change") or 0
        if abs(expected - change) > 0.001:
            r["quantity_mismatch"] = True
            r["document_quantity"] = expected
    # Qoldiq (harakatdan keyin) — ketma-ket yig'indi (eng eski harakatdan boshlab), eski yozuvlardagi xatolarni bartaraf etish
    if rows:
        chronological = list(reversed(rows))
        balance = 0.0
        for r in chronological:
            balance += r["quantity_change"]
            r["quantity_after"] = round(balance, 6)
        rows = list(reversed(chronological))
    current_stock = db.query(Stock).filter(
        Stock.warehouse_id == warehouse_id,
        Stock.product_id == product_id,
    ).all()
    current_qty = sum(float(s.quantity or 0) for s in current_stock)
    # Dona mahsulotlar uchun ko'rsatishda butun son (216, 6), boshqalar uchun 3 xona kasr
    _unit_str = ((getattr(product, "unit", None) and (product.unit.name or "") or "") + " " + (getattr(product, "unit", None) and (product.unit.code or "") or "")).lower()
    is_dona = product and "dona" in _unit_str
    return templates.TemplateResponse("reports/stock_source.html", {
        "request": request,
        "warehouse": warehouse,
        "product": product,
        "movements": rows,
        "current_qty": current_qty,
        "is_dona": is_dona,
        "page_title": "Qoldiq manbai",
        "current_user": current_user,
    })


@router.post("/stock/merge-duplicates")
async def report_stock_merge_duplicates(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Bir xil (ombor, mahsulot) uchun bitta Stock qatori qoldiradi, qolganlarini o'chirib yig'indini bitta qatorga yozadi (faqat admin)."""
    from collections import defaultdict
    all_stocks = db.query(Stock).all()
    by_key = defaultdict(list)
    for s in all_stocks:
        by_key[(s.warehouse_id, s.product_id)].append(s)
    merged = 0
    for key, group in by_key.items():
        if len(group) <= 1:
            continue
        total = sum(float(s.quantity or 0) for s in group)
        keep = group[0]
        keep.quantity = total
        if hasattr(keep, "updated_at"):
            keep.updated_at = datetime.now()
        for s in group[1:]:
            db.delete(s)
            merged += 1
    db.commit()
    return RedirectResponse(
        url=f"/reports/stock?merged={merged}",
        status_code=303,
    )


@router.post("/stock/recalculate-from-movements")
async def report_stock_recalculate_from_movements(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Stock qoldiqlarini StockMovement tarixidan qayta hisoblash (faqat admin). Eski ikki marta qo'shilgan xatolikni tuzatish uchun."""
    from collections import defaultdict
    # O'chirilgan ishlab chiqarish hujjatlariga tegishli "orfan" harakatlarni o'chirish (qoldiq to'g'ri tushadi)
    existing_production_ids = {r[0] for r in db.query(Production.id).all()}
    orphan_production_movements = db.query(StockMovement).filter(
        StockMovement.document_type == "Production",
        StockMovement.document_id.isnot(None),
    ).all()
    deleted_orphans = 0
    for m in orphan_production_movements:
        if m.document_id not in existing_production_ids:
            db.delete(m)
            deleted_orphans += 1
    if deleted_orphans:
        db.flush()
    # Tasdiqlangan qoldiq tuzatish hujjatlarini aniqlash (boshqa hujjat turlari hammasi hisobga olinadi)
    adj_ids = db.query(StockMovement.document_id).filter(
        StockMovement.document_type == "StockAdjustmentDoc",
        StockMovement.document_id.isnot(None),
    ).distinct().all()
    adj_ids = [r[0] for r in adj_ids if r[0]]
    confirmed_adj_ids = set()
    if adj_ids:
        for doc in db.query(StockAdjustmentDoc).filter(
            StockAdjustmentDoc.id.in_(adj_ids),
            StockAdjustmentDoc.status == "confirmed",
        ).all():
            confirmed_adj_ids.add(doc.id)
    # Har bir (warehouse_id, product_id) uchun harakatlarni created_at bo'yicha olamiz
    movements = (
        db.query(StockMovement)
        .filter(
            StockMovement.warehouse_id.isnot(None),
            StockMovement.product_id.isnot(None),
        )
        .order_by(StockMovement.created_at.asc())
        .all()
    )
    # (warehouse_id, product_id) -> ro'yxatda har bir hujjat (document_type, document_id) uchun bitta harakat (oldingi ikki marta qo'shilgan xatoni bartaraf etish)
    by_key = defaultdict(list)
    for m in movements:
        if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id not in confirmed_adj_ids:
            continue
        key = (m.warehouse_id, m.product_id)
        by_key[key].append(m)
    totals = {}
    for key, lst in by_key.items():
        seen_doc = set()
        s = 0.0
        for m in lst:
            doc_key = (m.document_type or "", m.document_id)
            if doc_key in seen_doc:
                continue
            seen_doc.add(doc_key)
            s += float(m.quantity_change or 0)
        totals[key] = s
    updated = 0
    created = 0
    for (wh_id, prod_id), qty in totals.items():
        qty = max(0.0, qty)
        stock = db.query(Stock).filter(
            Stock.warehouse_id == wh_id,
            Stock.product_id == prod_id,
        ).first()
        if stock:
            stock.quantity = qty
            if hasattr(stock, "updated_at"):
                stock.updated_at = datetime.now()
            updated += 1
        else:
            db.add(Stock(warehouse_id=wh_id, product_id=prod_id, quantity=qty))
            created += 1
    db.commit()
    from urllib.parse import quote
    msg_parts = [f"Stock qoldiqlari harakatlar tarixidan qayta hisoblandi: {updated} yangilandi, {created} yangi qator."]
    if deleted_orphans:
        msg_parts.append(f" O'chirilgan ishlab chiqarish hujjatlariga tegishli {deleted_orphans} ta harakat olib tashlandi.")
    msg = quote("".join(msg_parts))
    return RedirectResponse(url=f"/reports/stock?recalculated=1&msg={msg}", status_code=303)


@router.post("/stock/clear")
async def report_stock_clear(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Stock jadvalini to'liq tozalash — barcha qoldiq yozuvlarini o'chirish (faqat admin). StockMovement tarixi saqlanadi."""
    db.query(StockMovement).filter(StockMovement.stock_id.isnot(None)).update({StockMovement.stock_id: None}, synchronize_session=False)
    deleted = db.query(Stock).delete()
    db.commit()
    from urllib.parse import quote
    return RedirectResponse(
        url=f"/reports/stock?cleared={deleted}&msg=" + quote("Stock jadvali tozalandi. Qoldiq hisoboti endi bo'sh."),
        status_code=303,
    )


def _stock_report_filtered(db: Session, wh_id: int = None):
    """Stock jadvalidan faqat tasdiqlangan manba va qoldiq > 0 bo'lgan qatorlarni qaytaradi (hisobot va eksport uchun)."""
    q = (
        db.query(Stock)
        .join(Product, Stock.product_id == Product.id)
        .join(Warehouse, Stock.warehouse_id == Warehouse.id)
        .order_by(Warehouse.name, Product.name)
    )
    if wh_id:
        q = q.filter(Stock.warehouse_id == wh_id)
    rows = q.all()
    aggregated = {}
    for s in rows:
        key = (s.warehouse_id, s.product_id)
        if key not in aggregated:
            aggregated[key] = {"warehouse": s.warehouse, "product": s.product, "quantity": 0}
        aggregated[key]["quantity"] += float(s.quantity or 0)
    keys = list(aggregated.keys())
    allowed_keys = set()
    if keys:
        mov_q = db.query(StockMovement).filter(
            or_(*[and_(StockMovement.warehouse_id == k[0], StockMovement.product_id == k[1]) for k in keys])
        ).all()
        adj_doc_ids = {m.document_id for m in mov_q if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id}
        confirmed_adj_ids = set()
        if adj_doc_ids:
            confirmed_adj_ids = {
                d.id for d in db.query(StockAdjustmentDoc.id).filter(
                    StockAdjustmentDoc.id.in_(adj_doc_ids),
                    StockAdjustmentDoc.status == "confirmed",
                ).all()
            }
        for m in mov_q:
            key = (m.warehouse_id, m.product_id)
            if key not in aggregated:
                continue
            if (m.document_type or "") == "StockAdjustmentDoc":
                if m.document_id in confirmed_adj_ids:
                    allowed_keys.add(key)
            else:
                allowed_keys.add(key)
    if allowed_keys:
        aggregated = {k: v for k, v in aggregated.items() if k in allowed_keys}
    else:
        aggregated = {}
    aggregated = {k: v for k, v in aggregated.items() if float(v.get("quantity") or 0) > 0}
    return sorted(aggregated.values(), key=lambda x: ((x["warehouse"].name or "").lower(), (x["product"].name or "").lower()))


@router.get("/stock/export")
async def report_stock_export(
    warehouse_id: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    wh_id = None
    if warehouse_id is not None and str(warehouse_id).strip() != "":
        try:
            wh_id = int(warehouse_id)
        except (ValueError, TypeError):
            wh_id = None
    values = _stock_report_filtered(db, wh_id)
    stocks = [{"warehouse": v["warehouse"], "product": v["product"], "quantity": v["quantity"]} for v in values]
    wb = Workbook()
    ws = wb.active
    ws.title = "Qoldiq"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A1"] = "Qoldiq hisoboti"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.append(["Ombor", "Mahsulot", "Kod", "Qoldiq", "Minimal", "Narx", "Summa"])
    for c in range(1, 8):
        ws.cell(row=4, column=c).fill = header_fill
        ws.cell(row=4, column=c).font = Font(bold=True, color="FFFFFF")
    for s in stocks:
        p = s.product
        wh = s.warehouse
        min_s = getattr(p, "min_stock", 0) or 0
        price = getattr(p, "purchase_price", 0) or 0
        ws.append([
            wh.name if wh else "",
            p.name if p else "",
            (p.barcode or p.code or "") if p else "",
            float(s.quantity or 0),
            float(min_s),
            float(price),
            float((s.quantity or 0) * price),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=qoldiq_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/stock/andoza")
async def report_stock_andoza(current_user: User = Depends(require_auth)):
    """Qoldiqlar uchun Excel andoza (Tannarx va Sotuv narxi ixtiyoriy)."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    wb = Workbook()
    ws = wb.active
    ws.title = "Andoza"
    ws.append(["Ombor nomi (yoki kodi)", "Mahsulot nomi (yoki kodi)", "Qoldiq", "Tannarx (so'm)", "Sotuv narxi (so'm)"])
    ws.append(["Xom ashyo ombori", "Yong'oq", 30, "", ""])
    ws.append(["Xom ashyo ombori", "Bodom", 100, "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=qoldiqlar_andoza.xlsx"},
    )


@router.post("/stock/import")
async def report_stock_import(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Exceldan qoldiqlarni yuklash — hujjat qoralama holatida yaratiladi. doc_date bo'lsa shu sana ishlatiladi."""
    from urllib.parse import quote
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    file = form.get("file") or form.get("excel_file")
    if not file or not getattr(file, "filename", None):
        return RedirectResponse(url="/reports/stock?error=" + quote("Excel fayl tanlang"), status_code=303)
    try:
        contents = await file.read() if hasattr(file, "read") else (getattr(file, "file", None) and file.file.read() or b"")
    except Exception:
        contents = b""
    if not contents:
        return RedirectResponse(url="/reports/stock?error=" + quote("Fayl bo'sh"), status_code=303)
    doc_date_str = (form.get("doc_date") or "").strip()
    try:
        if doc_date_str:
            doc_date = datetime.strptime(doc_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=0, microsecond=0)
        else:
            doc_date = datetime.now()
    except ValueError:
        doc_date = datetime.now()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    items_data = []  # (product_id, warehouse_id, qty, tannarx, sotuv_narxi)
    for row in rows:
        if not row or (row[0] is None or row[0] == "") or (row[1] is None or row[1] == ""):
            continue
        wh_key = str(row[0] or "").strip()
        raw_prod = row[1]
        if raw_prod is not None and isinstance(raw_prod, (int, float)) and float(raw_prod) == int(float(raw_prod)):
            product_key = str(int(float(raw_prod)))
        else:
            product_key = str(raw_prod or "").strip()
        try:
            qty = float(row[2]) if row[2] is not None and row[2] != "" else 0
        except (TypeError, ValueError):
            qty = 0
        tannarx = 0.0
        sotuv_narxi = 0.0
        if len(row) > 3 and row[3] is not None and row[3] != "":
            try:
                tannarx = float(row[3])
            except (TypeError, ValueError):
                pass
        if len(row) > 4 and row[4] is not None and row[4] != "":
            try:
                sotuv_narxi = float(row[4])
            except (TypeError, ValueError):
                pass
        wh = db.query(Warehouse).filter(
            (func.lower(Warehouse.name) == wh_key.lower()) | (Warehouse.code == wh_key)
        ).first()
        product = db.query(Product).filter(
            (Product.code == product_key) | (Product.barcode == product_key)
        ).first()
        if not product and product_key:
            product = db.query(Product).filter(
                Product.name.isnot(None),
                func.lower(Product.name) == product_key.lower()
            ).first()
        if not wh or not product:
            continue
        if tannarx > 0:
            product.purchase_price = tannarx
        if sotuv_narxi > 0:
            product.sale_price = sotuv_narxi
        items_data.append((product.id, wh.id, qty, tannarx, sotuv_narxi))
    if not items_data:
        return RedirectResponse(
            url="/reports/stock?error=" + quote("Hech qanday to'g'ri qator topilmadi"),
            status_code=303,
        )
    doc_date_start = doc_date.replace(hour=0, minute=0, second=0, microsecond=0)
    doc_date_end = doc_date_start + timedelta(days=1)
    count = db.query(StockAdjustmentDoc).filter(
        StockAdjustmentDoc.date >= doc_date_start,
        StockAdjustmentDoc.date < doc_date_end,
    ).count()
    number = f"QLD-{doc_date.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"
    total_tannarx = sum(qty * cp for _, _, qty, cp, _ in items_data)
    total_sotuv = sum(qty * sp for _, _, qty, _, sp in items_data)
    doc = StockAdjustmentDoc(
        number=number,
        date=doc_date,
        user_id=current_user.id if current_user else None,
        status="draft",
        total_tannarx=total_tannarx,
        total_sotuv=total_sotuv,
    )
    db.add(doc)
    db.flush()
    for pid, wid, qty, cp, sp in items_data:
        db.add(StockAdjustmentDocItem(
            doc_id=doc.id,
            product_id=pid,
            warehouse_id=wid,
            quantity=qty,
            cost_price=cp,
            sale_price=sp,
        ))
    db.commit()
    return RedirectResponse(
        url=f"/qoldiqlar/tovar/hujjat/{doc.id}?from=import&msg=" + quote("Hujjat qoralama. Qoldiq hisobotida ko'rinishi uchun «Tasdiqlash» bosing."),
        status_code=303,
    )


@router.get("/production", response_class=HTMLResponse)
async def report_production(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Ishlab chiqarish hisoboti"""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    q = (
        db.query(Production)
        .filter(
            Production.date >= start_date,
            Production.date <= end_date + " 23:59:59",
        )
        .order_by(Production.date.desc())
    )
    productions = q.all()
    total_qty = sum(p.quantity for p in productions if p.status == "completed")
    return templates.TemplateResponse("reports/production.html", {
        "request": request,
        "productions": productions,
        "total_qty": total_qty,
        "start_date": start_date,
        "end_date": end_date,
        "page_title": "Ishlab chiqarish hisoboti",
        "current_user": current_user,
    })


@router.get("/employees", response_class=HTMLResponse)
async def report_employees(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    allowed = get_allowed_report_types(current_user)
    if "employees" not in allowed and current_user.role != "admin":
        return RedirectResponse(url="/reports", status_code=303)
    employees = db.query(Employee).order_by(Employee.full_name).all()
    return templates.TemplateResponse("reports/employees.html", {
        "request": request,
        "employees": employees,
        "page_title": "Xodimlar hisoboti",
        "current_user": current_user,
    })


@router.get("/debts", response_class=HTMLResponse)
async def report_debts(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    debtors = db.query(Partner).filter(Partner.balance != 0).all()
    total_debt = sum(p.balance for p in debtors if p.balance > 0)
    total_credit = sum(abs(p.balance) for p in debtors if p.balance < 0)
    return templates.TemplateResponse("reports/debts.html", {
        "request": request,
        "debtors": debtors,
        "total_debt": total_debt,
        "total_credit": total_credit,
        "page_title": "Qarzdorlik hisoboti",
        "current_user": current_user,
    })


@router.get("/debts/export")
async def report_debts_export(db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    debtors = db.query(Partner).filter(Partner.balance != 0).order_by(Partner.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Qarzdorlik"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A1"] = "Qarzdorlik hisoboti"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.append(["Kod", "Mijoz", "Telefon", "Balans (qarz +)", "Kredit limiti"])
    for c in range(1, 6):
        ws.cell(row=4, column=c).fill = header_fill
        ws.cell(row=4, column=c).font = Font(bold=True, color="FFFFFF")
    for p in debtors:
        ws.append([
            p.code or "",
            p.name or "",
            p.phone or "",
            float(p.balance or 0),
            float(p.credit_limit or 0),
        ])
    total = sum(p.balance for p in debtors if (p.balance or 0) > 0)
    ws.append(["", "", "JAMI QARZ:", total, ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=qarzdorlik_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


def _build_partner_movements(db: Session, partner_id: int, date_from: datetime, date_to: datetime, period_only: bool):
    """
    Kontragent uchun harakatlar ro'yxati (1C uslubida).
    period_only=True: faqat [date_from, date_to] oralig'idagi qatorlar.
    Qaytadi: (rows, opening_debit, opening_credit) yoki period_only=True bo'lsa (rows, 0, 0).
    Balans: Debit = kontragent bizga qarzdor (sotuv), Credit = to'lov/xarid/qaytarish.
    """
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        return [], 0.0, 0.0
    date_from_start = date_from.replace(hour=0, minute=0, second=0, microsecond=0)
    date_to_end = date_to.replace(hour=23, minute=59, second=59, microsecond=999999)
    rows = []

    # Sotuvlar (debit) va qaytarishlar (credit)
    q_orders = db.query(Order).filter(
        Order.partner_id == partner_id,
        Order.type.in_(["sale", "return_sale"]),
    )
    if period_only:
        q_orders = q_orders.filter(Order.date >= date_from_start, Order.date <= date_to_end)
    for o in q_orders.order_by(Order.date):
        debit = float(o.total or 0) if o.type == "sale" else 0.0
        credit = float(o.total or 0) if o.type == "return_sale" else 0.0
        doc_label = f"{'Sotuv' if o.type == 'sale' else 'Qaytarish'} {o.number or ''} {o.date.strftime('%d.%m.%Y %H:%M') if o.date else ''}".strip()
        rows.append({
            "date": o.date,
            "doc_type": "Sotuv" if o.type == "sale" else "Qaytarish",
            "doc_number": o.number or "",
            "doc_label": doc_label,
            "doc_url": f"/sales/edit/{o.id}",
            "debit": debit,
            "credit": credit,
        })

    # To'lovlar: income = credit (ular bizga to'ladı), expense = debit (biz ularga to'ladık) — faqat tasdiqlangan
    q_payments = db.query(Payment).filter(Payment.partner_id == partner_id)
    if hasattr(Payment, "status"):
        q_payments = q_payments.filter(or_(Payment.status == "confirmed", Payment.status == None))
    if period_only:
        q_payments = q_payments.filter(Payment.date >= date_from_start, Payment.date <= date_to_end)
    for p in q_payments.order_by(Payment.date):
        if p.type == "income":
            doc_label = f"To'lov (kirim) {p.number or ''} {p.date.strftime('%d.%m.%Y %H:%M') if p.date else ''}".strip()
            rows.append({
                "date": p.date,
                "doc_type": "To'lov (kirim)",
                "doc_number": p.number or "",
                "doc_label": doc_label,
                "doc_url": f"/finance/payment/{p.id}/edit",
                "debit": 0.0,
                "credit": float(p.amount or 0),
            })
        else:
            doc_label = f"To'lov (chiqim) {p.number or ''} {p.date.strftime('%d.%m.%Y %H:%M') if p.date else ''}".strip()
            rows.append({
                "date": p.date,
                "doc_type": "To'lov (chiqim)",
                "doc_number": p.number or "",
                "doc_label": doc_label,
                "doc_url": f"/finance/payment/{p.id}/edit",
                "debit": float(p.amount or 0),
                "credit": 0.0,
            })

    # Xaridlar (biz yetkazuvchiga qarzdormiz — credit)
    q_purchases = db.query(Purchase).filter(Purchase.partner_id == partner_id)
    if period_only:
        q_purchases = q_purchases.filter(Purchase.date >= date_from_start, Purchase.date <= date_to_end)
    for p in q_purchases.order_by(Purchase.date):
        total_val = float((p.total or 0) + (p.total_expenses or 0))
        doc_label = f"Tovarlar kirimi (xarid) {p.number or ''} {p.date.strftime('%d.%m.%Y %H:%M') if p.date else ''}".strip()
        rows.append({
            "date": p.date,
            "doc_type": "Xarid",
            "doc_number": p.number or "",
            "doc_label": doc_label,
            "doc_url": f"/purchases/edit/{p.id}",
            "debit": 0.0,
            "credit": total_val,
        })

    rows.sort(key=lambda r: r["date"])
    opening_debit = 0.0
    opening_credit = 0.0
    if not period_only:
        # Opening = barcha harakatlar perioddan oldin
        for r in rows:
            if r["date"] < date_from_start:
                opening_debit += r["debit"]
                opening_credit += r["credit"]
        rows = [r for r in rows if date_from_start <= r["date"] <= date_to_end]
    return rows, opening_debit, opening_credit


@router.get("/partner-reconciliation", response_class=HTMLResponse)
async def report_partner_reconciliation(
    request: Request,
    partner_id: int = None,
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragentlar hisob-kitobini solishtirish hisoboti (1C uslubida)."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    allowed = get_allowed_report_types(current_user)
    if "partner_reconciliation" not in allowed:
        return RedirectResponse(url="/reports", status_code=303)
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    today = datetime.now()
    if not date_from:
        date_from = (today.replace(day=1)).strftime("%Y-%m-%d")
    if not date_to:
        date_to = today.strftime("%Y-%m-%d")
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d")
        dt_to = datetime.strptime(date_to, "%Y-%m-%d")
    except (ValueError, TypeError):
        dt_from = today.replace(day=1)
        dt_to = today
    rows = []
    opening_debit = opening_credit = 0.0
    total_debit = total_credit = 0.0
    partner_obj = None
    products_purchased = []  # kontragentdan xarid qilingan mahsulotlar (analitika)
    products_sold = []       # kontragentga sotilgan mahsulotlar (analitika)
    if partner_id:
        partner_obj = db.query(Partner).filter(Partner.id == partner_id).first()
        if partner_obj:
            rows, opening_debit, opening_credit = _build_partner_movements(db, partner_id, dt_from, dt_to, period_only=False)
            total_debit = sum(r["debit"] for r in rows)
            total_credit = sum(r["credit"] for r in rows)
            date_from_start = dt_from.replace(hour=0, minute=0, second=0, microsecond=0)
            date_to_end = dt_to.replace(hour=23, minute=59, second=59, microsecond=999999)
            # Kontragentdan xarid qilingan mahsulotlar (davr bo'yicha)
            purchases_in_period = (
                db.query(PurchaseItem, Product)
                .join(Purchase, PurchaseItem.purchase_id == Purchase.id)
                .join(Product, PurchaseItem.product_id == Product.id)
                .filter(
                    Purchase.partner_id == partner_id,
                    Purchase.date >= date_from_start,
                    Purchase.date <= date_to_end,
                )
            ).all()
            by_product_purchase = {}
            for pi, prod in purchases_in_period:
                key = prod.id
                if key not in by_product_purchase:
                    by_product_purchase[key] = {"product_name": prod.name or "", "product_code": prod.code or "", "quantity": 0.0, "total": 0.0}
                by_product_purchase[key]["quantity"] += float(pi.quantity or 0)
                by_product_purchase[key]["total"] += float(pi.total or 0)
            products_purchased = sorted(by_product_purchase.values(), key=lambda x: -x["total"])
            # Kontragentga sotilgan mahsulotlar (davr bo'yicha)
            orders_in_period = (
                db.query(OrderItem, Product)
                .join(Order, OrderItem.order_id == Order.id)
                .join(Product, OrderItem.product_id == Product.id)
                .filter(
                    Order.partner_id == partner_id,
                    Order.type == "sale",
                    Order.date >= date_from_start,
                    Order.date <= date_to_end,
                )
            ).all()
            by_product_sale = {}
            for oi, prod in orders_in_period:
                key = prod.id
                if key not in by_product_sale:
                    by_product_sale[key] = {"product_name": prod.name or "", "product_code": prod.code or "", "quantity": 0.0, "total": 0.0}
                by_product_sale[key]["quantity"] += float(oi.quantity or 0)
                by_product_sale[key]["total"] += float(oi.total or 0)
            products_sold = sorted(by_product_sale.values(), key=lambda x: -x["total"])
    opening_balance = opening_debit - opening_credit
    closing_balance = opening_balance + total_debit - total_credit
    return templates.TemplateResponse("reports/partner_reconciliation.html", {
        "request": request,
        "partners": partners,
        "partner_id": partner_id,
        "partner": partner_obj,
        "date_from": date_from,
        "date_to": date_to,
        "rows": rows,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "opening_debit": opening_debit,
        "opening_credit": opening_credit,
        "products_purchased": products_purchased,
        "products_sold": products_sold,
        "page_title": "Kontragentlar hisob-kitobini solishtirish",
        "current_user": current_user,
    })


@router.get("/partner-reconciliation/export")
async def report_partner_reconciliation_export(
    partner_id: int = None,
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent solishtirish hisobotini Excelga eksport."""
    if not current_user:
        return RedirectResponse(url="/reports", status_code=303)
    if not partner_id:
        return RedirectResponse(url="/reports/partner-reconciliation?error=partner", status_code=303)
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        return RedirectResponse(url="/reports/partner-reconciliation", status_code=303)
    today = datetime.now()
    if not date_from:
        date_from = (today.replace(day=1)).strftime("%Y-%m-%d")
    if not date_to:
        date_to = today.strftime("%Y-%m-%d")
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d")
        dt_to = datetime.strptime(date_to, "%Y-%m-%d")
    except (ValueError, TypeError):
        dt_from = today.replace(day=1)
        dt_to = today
    rows, opening_debit, opening_credit = _build_partner_movements(db, partner_id, dt_from, dt_to, period_only=False)
    total_debit = sum(r["debit"] for r in rows)
    total_credit = sum(r["credit"] for r in rows)
    opening_balance = opening_debit - opening_credit
    closing_balance = opening_balance + total_debit - total_credit

    # Mahsulotlar bo'yicha analitika (veb sahifadagi kabi)
    date_from_start = dt_from.replace(hour=0, minute=0, second=0, microsecond=0)
    date_to_end = dt_to.replace(hour=23, minute=59, second=59, microsecond=999999)
    by_product_purchase = {}
    for pi, prod in (
        db.query(PurchaseItem, Product)
        .join(Purchase, PurchaseItem.purchase_id == Purchase.id)
        .join(Product, PurchaseItem.product_id == Product.id)
        .filter(
            Purchase.partner_id == partner_id,
            Purchase.date >= date_from_start,
            Purchase.date <= date_to_end,
        )
    ).all():
        key = prod.id
        if key not in by_product_purchase:
            by_product_purchase[key] = {"product_name": prod.name or "", "product_code": prod.code or "", "quantity": 0.0, "total": 0.0}
        by_product_purchase[key]["quantity"] += float(pi.quantity or 0)
        by_product_purchase[key]["total"] += float(pi.total or 0)
    products_purchased = sorted(by_product_purchase.values(), key=lambda x: -x["total"])
    by_product_sale = {}
    for oi, prod in (
        db.query(OrderItem, Product)
        .join(Order, OrderItem.order_id == Order.id)
        .join(Product, OrderItem.product_id == Product.id)
        .filter(
            Order.partner_id == partner_id,
            Order.type == "sale",
            Order.date >= date_from_start,
            Order.date <= date_to_end,
        )
    ).all():
        key = prod.id
        if key not in by_product_sale:
            by_product_sale[key] = {"product_name": prod.name or "", "product_code": prod.code or "", "quantity": 0.0, "total": 0.0}
        by_product_sale[key]["quantity"] += float(oi.quantity or 0)
        by_product_sale[key]["total"] += float(oi.total or 0)
    products_sold = sorted(by_product_sale.values(), key=lambda x: -x["total"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Solishtirish"
    ws["A1"] = "Hisob kitoblarni solishtirish"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"TOTLI HOLVA va {partner.name or ''}"
    ws["A3"] = f"Davr: {date_from} — {date_to}"
    # Analitika bloki (veb sahifadagi summary ga mos)
    ws["A4"] = "Davlati qoldiq (" + date_from + "):"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = opening_balance
    ws["A5"] = "Davr debet:"
    ws["B5"] = total_debit
    ws["A6"] = "Davr kredit:"
    ws["B6"] = total_credit
    ws["A7"] = "Yakuniy qoldiq (" + date_to + "):"
    ws["A7"].font = Font(bold=True)
    ws["B7"] = closing_balance
    ws["A8"] = "Bizning foydamizga (kontragent qarzdor):"
    ws["B8"] = closing_balance if closing_balance > 0 else 0
    ws["A9"] = "Kontragent foydasiga (biz qarzdormiz):"
    ws["B9"] = -closing_balance if closing_balance < 0 else 0
    ws.append([])
    table_start = 11
    headers = ["Hujjatlar", "TOTLI HOLVA DT (so'm)", "TOTLI HOLVA KT (so'm)", f"{partner.name or 'Kontragent'} DT (so'm)", f"{partner.name or 'Kontragent'} KT (so'm)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=table_start, column=c, value=h)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    row_num = table_start + 1
    ws.cell(row=row_num, column=1, value=f"{date_from} sanaga qoldiq")
    ws.cell(row=row_num, column=2, value=opening_balance if opening_balance > 0 else 0)
    ws.cell(row=row_num, column=3, value=-opening_balance if opening_balance < 0 else 0)
    ws.cell(row=row_num, column=4, value=-opening_balance if opening_balance < 0 else 0)
    ws.cell(row=row_num, column=5, value=opening_balance if opening_balance > 0 else 0)
    row_num += 1
    for r in rows:
        ws.cell(row=row_num, column=1, value=r.get("doc_label") or f"{r.get('doc_type', '')} {r.get('doc_number', '')}")
        ws.cell(row=row_num, column=2, value=r["debit"] if r["debit"] else 0)
        ws.cell(row=row_num, column=3, value=r["credit"] if r["credit"] else 0)
        ws.cell(row=row_num, column=4, value=r["credit"] if r["credit"] else 0)
        ws.cell(row=row_num, column=5, value=r["debit"] if r["debit"] else 0)
        row_num += 1
    ws.cell(row=row_num, column=1, value="Jami davr:")
    ws.cell(row=row_num, column=2, value=total_debit)
    ws.cell(row=row_num, column=3, value=total_credit)
    ws.cell(row=row_num, column=4, value=total_credit)
    ws.cell(row=row_num, column=5, value=total_debit)
    row_num += 1
    ws.cell(row=row_num, column=1, value=f"{date_to} sanaga qoldiq")
    ws.cell(row=row_num, column=2, value=closing_balance if closing_balance > 0 else 0)
    ws.cell(row=row_num, column=3, value=-closing_balance if closing_balance < 0 else 0)
    ws.cell(row=row_num, column=4, value=-closing_balance if closing_balance < 0 else 0)
    ws.cell(row=row_num, column=5, value=closing_balance if closing_balance > 0 else 0)
    ws.column_dimensions["A"].width = 52
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 18

    # Varaq: Kontragentdan xarid qilingan mahsulotlar
    ws_purchase = wb.create_sheet("Xarid qilingan", 1)
    ws_purchase["A1"] = "Kontragentdan xarid qilingan mahsulotlar"
    ws_purchase["A1"].font = Font(bold=True, size=12)
    ws_purchase["A2"] = f"Kontragent: {partner.name or ''}  |  Davr: {date_from} — {date_to}"
    for c, h in enumerate(["Mahsulot", "Kod", "Miqdor", "Summa (so'm)"], 1):
        cell = ws_purchase.cell(row=4, column=c, value=h)
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    for i, p in enumerate(products_purchased, 5):
        ws_purchase.cell(row=i, column=1, value=p["product_name"])
        ws_purchase.cell(row=i, column=2, value=p["product_code"])
        ws_purchase.cell(row=i, column=3, value=p["quantity"])
        ws_purchase.cell(row=i, column=4, value=p["total"])
    if not products_purchased:
        ws_purchase.cell(row=5, column=1, value="Davrda xarid qilinmagan.")

    # Varaq: Kontragentga sotilgan mahsulotlar
    ws_sale = wb.create_sheet("Sotilgan", 2)
    ws_sale["A1"] = "Kontragentga sotilgan mahsulotlar"
    ws_sale["A1"].font = Font(bold=True, size=12)
    ws_sale["A2"] = f"Kontragent: {partner.name or ''}  |  Davr: {date_from} — {date_to}"
    for c, h in enumerate(["Mahsulot", "Kod", "Miqdor", "Summa (so'm)"], 1):
        cell = ws_sale.cell(row=4, column=c, value=h)
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    for i, p in enumerate(products_sold, 5):
        ws_sale.cell(row=i, column=1, value=p["product_name"])
        ws_sale.cell(row=i, column=2, value=p["product_code"])
        ws_sale.cell(row=i, column=3, value=p["quantity"])
        ws_sale.cell(row=i, column=4, value=p["total"])
    if not products_sold:
        ws_sale.cell(row=5, column=1, value="Davrda sotuv bo'lmagan.")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"kontragent_solishtirish_{partner.id}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
