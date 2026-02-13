"""
Dashboard Export Routes
Excel va PDF export funksiyalari
"""

from fastapi import Request, Depends
from fastapi.responses import StreamingResponse, RedirectResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.models.database import (
    get_db, User, Order, OrderItem, Agent, Stock, Product, Partner
)
from app.utils.auth import get_user_from_token


async def export_executive_dashboard(request: Request, db: Session):
    """Export Executive Dashboard to Excel"""
    
    # Get user from session cookie
    session_token = request.cookies.get("session_token")
    if not session_token:
        return RedirectResponse(url="/login", status_code=303)
    
    user_data = get_user_from_token(session_token)
    if not user_data:
        return RedirectResponse(url="/login", status_code=303)
    
    user = db.query(User).filter(User.id == user_data["user_id"]).first()
    if not user or not user.is_active:
        return RedirectResponse(url="/login", status_code=303)
    
    # Get data
    today = datetime.now().date()
    week_ago = today - timedelta(days=7)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rahbariyat Hisoboti"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws['A1'] = 'TOTLI HOLVA - Rahbariyat Hisoboti'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f'Sana: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    
    # Today's sales
    today_sales = db.query(func.sum(Order.total)).filter(
        func.date(Order.created_at) == today,
        Order.status == 'completed'
    ).scalar() or 0
    
    yesterday_sales = db.query(func.sum(Order.total)).filter(
        func.date(Order.created_at) == today - timedelta(days=1),
        Order.status == 'completed'
    ).scalar() or 0
    
    sales_growth = 0
    if yesterday_sales > 0:
        sales_growth = ((today_sales - yesterday_sales) / yesterday_sales) * 100
    
    # KPI Section
    ws['A4'] = 'ASOSIY KO\'RSATKICHLAR'
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws.merge_cells('A4:B4')
    
    ws['A5'] = 'Bugungi savdo'
    ws['B5'] = f"{today_sales:,.0f} so'm"
    ws['A6'] = 'O\'sish'
    ws['B6'] = f"{sales_growth:.1f}%"
    
    today_orders = db.query(func.count(Order.id)).filter(
        func.date(Order.created_at) == today
    ).scalar() or 0
    
    ws['A7'] = 'Bugungi buyurtmalar'
    ws['B7'] = today_orders
    
    active_agents = db.query(func.count(Agent.id)).filter(
        Agent.is_active == True
    ).scalar() or 0
    
    ws['A8'] = 'Faol agentlar'
    ws['B8'] = active_agents
    
    # Top Products
    ws['A10'] = 'TOP 5 MAHSULOTLAR'
    ws['A10'].font = header_font
    ws['A10'].fill = header_fill
    ws.merge_cells('A10:C10')
    
    ws['A11'] = '#'
    ws['B11'] = 'Mahsulot'
    ws['C11'] = 'Miqdor'
    for cell in ['A11', 'B11', 'C11']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    top_products = db.query(
        Product.name,
        func.sum(OrderItem.quantity).label('total_qty')
    ).join(
        OrderItem, Product.id == OrderItem.product_id
    ).join(
        Order, OrderItem.order_id == Order.id
    ).filter(
        func.date(Order.created_at) >= week_ago,
        Order.status == 'completed'
    ).group_by(Product.id, Product.name).order_by(
        func.sum(OrderItem.quantity).desc()
    ).limit(5).all()
    
    row = 12
    for i, (name, qty) in enumerate(top_products, 1):
        ws[f'A{row}'] = i
        ws[f'B{row}'] = name
        ws[f'C{row}'] = int(qty)
        row += 1
    
    # Top Agents
    ws[f'A{row+1}'] = 'TOP 5 AGENTLAR'
    ws[f'A{row+1}'].font = header_font
    ws[f'A{row+1}'].fill = header_fill
    ws.merge_cells(f'A{row+1}:D{row+1}')
    
    row += 2
    ws[f'A{row}'] = '#'
    ws[f'B{row}'] = 'Agent'
    ws[f'C{row}'] = 'Savdo'
    ws[f'D{row}'] = 'Buyurtmalar'
    for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    top_agents = db.query(
        Agent.full_name,
        func.sum(Order.total).label('total_sales'),
        func.count(Order.id).label('order_count')
    ).join(
        Order, Agent.id == Order.partner_id
    ).filter(
        func.date(Order.created_at) >= week_ago,
        Order.status == 'completed'
    ).group_by(Agent.id, Agent.full_name).order_by(
        func.sum(Order.total).desc()
    ).limit(5).all()
    
    row += 1
    for i, agent in enumerate(top_agents, 1):
        ws[f'A{row}'] = i
        ws[f'B{row}'] = agent.full_name
        ws[f'C{row}'] = f"{agent.total_sales:,.0f} so'm"
        ws[f'D{row}'] = agent.order_count
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"rahbariyat_hisobot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
